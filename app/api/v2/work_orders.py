from fastapi import APIRouter, HTTPException, status, Query, Request
from fastapi.encoders import jsonable_encoder
from sqlalchemy import select, func, cast, String, text, and_, or_, desc
from sqlalchemy.orm import aliased
from typing import Optional, List
from datetime import datetime, date as date_type
from pydantic import BaseModel, Field
import uuid
import logging
import traceback

from app.api.deps import DbSession, CurrentUser, EntityCtx
from app.models.work_order import WorkOrder
from app.models.work_order_audit import WorkOrderAuditLog
from app.models.customer import Customer
from app.models.technician import Technician
from app.services.commission_service import auto_create_commission
from app.services.cache_service import get_cache_service, TTL
from app.schemas.work_order import (
    WorkOrderCreate,
    WorkOrderUpdate,
    WorkOrderResponse,
    WorkOrderListResponse,
    WorkOrderCursorResponse,
    WorkOrderAuditLogResponse,
)
from app.schemas.pagination import decode_cursor, encode_cursor
from app.services.websocket_manager import manager

logger = logging.getLogger(__name__)

router = APIRouter()

# Alias for billing customer JOIN (same table, different FK)
BillingCustomer = aliased(Customer, name="billing_customer")

# PostgreSQL ENUM fields that need explicit type casting
ENUM_FIELDS = {"status", "job_type", "priority"}

# Real estate inspection reports are always sent to Doug
REAL_ESTATE_INSPECTION_RECIPIENT = "Doug@macseptic.com"


async def _send_real_estate_inspection_report(wo: WorkOrder, db) -> None:
    """Auto-email the inspection report PDF to Doug when a real_estate_inspection is completed."""
    if str(wo.job_type) != "real_estate_inspection":
        return

    checklist = wo.checklist or {}
    inspection = checklist.get("inspection", {})
    if not inspection:
        logger.info(f"[RE-INSPECTION] No inspection data for WO {wo.id}, skipping email")
        return

    try:
        # Get customer info
        customer_name = "Valued Customer"
        customer_address = ""
        if wo.customer_id:
            cust_result = await db.execute(select(Customer).where(Customer.id == wo.customer_id))
            cust = cust_result.scalars().first()
            if cust:
                customer_name = f"{cust.first_name or ''} {cust.last_name or ''}".strip() or "Valued Customer"
                addr_parts = [p for p in [cust.address_line1, cust.city, cust.state, cust.postal_code] if p]
                customer_address = ", ".join(addr_parts)

        # Generate PDF
        from app.services.inspection_pdf import generate_inspection_pdf, pdf_to_base64
        pdf_bytes = generate_inspection_pdf(
            customer_name=customer_name,
            customer_address=customer_address,
            inspection_data=inspection,
            work_order_id=str(wo.id),
            job_type="Real Estate Inspection",
            scheduled_date=str(wo.scheduled_date) if wo.scheduled_date else None,
        )
        pdf_b64 = pdf_to_base64(pdf_bytes)

        # Build email
        summary = inspection.get("summary", {})
        condition = summary.get("overall_condition", "N/A")
        condition_label = "Good" if condition == "good" else "Needs Attention" if condition == "fair" else "Needs Repair" if condition in ("poor", "critical") else condition.title()

        wo_number = wo.work_order_number or str(wo.id)[:8]
        addr_parts = [wo.service_address_line1, wo.service_city, wo.service_state]
        service_addr = ", ".join(p for p in addr_parts if p) or "N/A"

        subject = f"Real Estate Inspection Complete — {customer_name} — {condition_label}"
        plain_text = (
            f"Real Estate Inspection Report\n\n"
            f"Work Order: {wo_number}\n"
            f"Customer: {customer_name}\n"
            f"Address: {service_addr}\n"
            f"Overall Condition: {condition_label}\n\n"
            f"The full PDF report is attached."
        )
        html_body = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto">
          <div style="background:#1e3a5f;color:white;padding:20px;text-align:center;border-radius:8px 8px 0 0">
            <h2 style="margin:0;font-size:18px">Real Estate Inspection Report</h2>
          </div>
          <div style="padding:20px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 8px 8px">
            <table style="width:100%;border-collapse:collapse;margin-bottom:16px">
              <tr><td style="padding:6px 0;color:#6b7280;font-size:13px">Work Order</td><td style="padding:6px 0;font-weight:600">{wo_number}</td></tr>
              <tr><td style="padding:6px 0;color:#6b7280;font-size:13px">Customer</td><td style="padding:6px 0;font-weight:600">{customer_name}</td></tr>
              <tr><td style="padding:6px 0;color:#6b7280;font-size:13px">Address</td><td style="padding:6px 0;font-weight:600">{service_addr}</td></tr>
              <tr><td style="padding:6px 0;color:#6b7280;font-size:13px">Condition</td><td style="padding:6px 0;font-weight:700;color:{'#22c55e' if condition == 'good' else '#f59e0b' if condition == 'fair' else '#ef4444'}">{condition_label}</td></tr>
            </table>
            <p style="margin:16px 0 0;padding:12px;background:#eff6ff;border-radius:6px;text-align:center;font-size:14px;color:#1e40af">
              <strong>Full PDF report is attached.</strong>
            </p>
          </div>
        </div>
        """

        from app.services.email_service import EmailService
        email_svc = EmailService()
        if not email_svc.is_configured:
            logger.warning("[RE-INSPECTION] Email service not configured, skipping report email")
            return

        result = await email_svc.send_email(
            to=REAL_ESTATE_INSPECTION_RECIPIENT,
            subject=subject,
            body=plain_text,
            html_body=html_body,
            attachments=[{
                "content": pdf_b64,
                "name": f"RE-Inspection-{wo_number}.pdf",
            }],
        )
        if result.get("success"):
            logger.info(f"[RE-INSPECTION] Report emailed to {REAL_ESTATE_INSPECTION_RECIPIENT} for WO {wo.id}")
        else:
            logger.warning(f"[RE-INSPECTION] Email failed for WO {wo.id}: {result.get('error')}")

    except Exception as e:
        logger.error(f"[RE-INSPECTION] Failed to send report email for WO {wo.id}: {e}")


@router.get("/inspection-letters/queue")
async def get_inspection_letter_queue(db: DbSession, current_user: CurrentUser):
    """Get all RE inspection work orders with their AI letter status.

    This is an admin/office endpoint (not employee portal) so Doug can
    access it from the main CRM sidebar.
    """
    try:
        from sqlalchemy.orm import selectinload
        from sqlalchemy import nullslast

        result = await db.execute(
            select(WorkOrder)
            .options(selectinload(WorkOrder.customer))
            .where(cast(WorkOrder.job_type, String) == "real_estate_inspection")
            .order_by(nullslast(WorkOrder.scheduled_date.desc()))
            .limit(100)
        )
        work_orders = result.scalars().all()

        items = []
        for wo in work_orders:
            checklist = wo.checklist if isinstance(wo.checklist, dict) else {}
            inspection = checklist.get("inspection") or {}
            ai_letter = inspection.get("ai_letter") or {}
            summary = inspection.get("summary") or {}

            customer_name = "Unknown"
            customer_email = None
            if wo.customer:
                customer_name = f"{wo.customer.first_name or ''} {wo.customer.last_name or ''}".strip() or "Unknown"
                customer_email = wo.customer.email

            address = ""
            if wo.service_address_line1:
                parts = [wo.service_address_line1, wo.service_city, wo.service_state, wo.service_postal_code]
                address = ", ".join(p for p in parts if p)
            elif wo.customer:
                parts = [wo.customer.address_line1, wo.customer.city, wo.customer.state, wo.customer.postal_code]
                address = ", ".join(p for p in parts if p)

            letter_status = ai_letter.get("status", "none")
            has_inspection_data = bool(
                inspection.get("steps")
                or inspection.get("client")
                or inspection.get("findings")
            )

            items.append({
                "id": str(wo.id),
                "work_order_number": wo.work_order_number,
                "customer_name": customer_name,
                "customer_email": customer_email,
                "address": address,
                "scheduled_date": wo.scheduled_date.isoformat() if wo.scheduled_date else None,
                "status": str(wo.status) if wo.status else None,
                "letter_status": letter_status,
                "has_inspection_data": has_inspection_data,
                "overall_condition": summary.get("overall_condition"),
                "sent_at": ai_letter.get("sent_at"),
                "sent_to": ai_letter.get("sent_to"),
            })

        return {"items": items, "total": len(items)}
    except Exception as e:
        logger.error(f"[INSPECTION-LETTERS] Queue endpoint error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/inspection-letters/debug-row")
async def debug_form_row(
    row_index: int = 0,
    db: DbSession = None,
    current_user: CurrentUser = None,
):
    """Dump a full row from the MS Forms Excel so we can see the column layout."""
    from app.services.ms365_forms_sync_service import MS365FormsSyncService
    import urllib.parse

    if not MS365FormsSyncService.is_configured():
        return {"error": "MS365 not configured"}

    try:
        site_id = await MS365FormsSyncService._get_site_id()
        drive_id, item_id = await MS365FormsSyncService._get_drive_item_id(site_id)

        # Read the FULL used range including the header row
        data = await MS365FormsSyncService.graph_get(
            f"/drives/{drive_id}/items/{item_id}/workbook/worksheets/{urllib.parse.quote(await MS365FormsSyncService._get_worksheet_name(drive_id, item_id), safe='')}/usedRange"
        )
        all_rows = data.get("values", [])
        if not all_rows:
            return {"error": "No rows"}

        headers = all_rows[0]
        sample_row = all_rows[row_index + 1] if row_index + 1 < len(all_rows) else all_rows[1]

        return {
            "total_rows": len(all_rows),
            "headers": [{"idx": i, "name": h} for i, h in enumerate(headers)],
            "sample_row": [{"idx": i, "value": str(v)[:200] if v is not None else ""} for i, v in enumerate(sample_row)],
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "trace": traceback.format_exc()[:2000]}


@router.get("/inspection-letters/debug-find")
async def debug_find_in_forms(
    name: str = "",
    address: str = "",
    db: DbSession = None,
    current_user: CurrentUser = None,
):
    """Debug: search the Forms workbook for a customer name/address match."""
    from app.services.ms365_forms_sync_service import MS365FormsSyncService, COL, _safe_get

    if not MS365FormsSyncService.is_configured():
        return {"error": "MS365 not configured"}

    try:
        site_id = await MS365FormsSyncService._get_site_id()
        drive_id, item_id = await MS365FormsSyncService._get_drive_item_id(site_id)
        sheet_name = await MS365FormsSyncService._get_worksheet_name(drive_id, item_id)
        rows = await MS365FormsSyncService._read_rows(drive_id, item_id, sheet_name)

        name_lower = name.lower()
        addr_lower = address.lower()

        matches = []
        all_names = []
        for i, row in enumerate(rows):
            row_name = _safe_get(row, COL["client_name"])
            row_addr = _safe_get(row, COL["client_address"])
            all_names.append(f"{row_name} @ {row_addr}")

            # Simple substring match in either direction
            if (name_lower and (name_lower in row_name.lower() or row_name.lower() in name_lower)) or \
               (addr_lower and (addr_lower[:20] in row_addr.lower() or row_addr.lower()[:20] in addr_lower)):
                matches.append({
                    "row": i + 2,
                    "name": row_name,
                    "address": row_addr,
                    "full_row": [str(c)[:100] for c in row],
                })

        # Also try the actual find function
        find_result = await MS365FormsSyncService.find_inspection_by_address_or_name(
            customer_name=name,
            address=address,
        )

        return {
            "total_rows": len(rows),
            "search_name": name,
            "search_address": address,
            "matches": matches[:5],
            "find_function_result": bool(find_result),
            "find_function_data": find_result if find_result else None,
            "all_rows_sample": all_names[:20],
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "trace": traceback.format_exc()[:2000]}


@router.post("/inspection-letters/upload-forms")
async def upload_forms_excel(
    request: Request,
    db: DbSession,
    current_user: CurrentUser,
):
    """Accept a manually-uploaded MS Forms Excel export and cache the rows.

    Workaround for when Azure app lacks SharePoint file permissions.
    Parses the uploaded .xlsx, extracts all rows, and stores them in
    the cache service for subsequent letter generation.
    """
    try:
        import io
        from openpyxl import load_workbook
        from app.services.cache_service import get_cache_service, TTL

        # Get the uploaded file from multipart form
        form = await request.form()
        upload = form.get("file")
        if upload is None or not hasattr(upload, "read"):
            raise HTTPException(status_code=400, detail="No file uploaded")

        content = await upload.read()
        if not content:
            raise HTTPException(status_code=400, detail="File is empty")

        # Parse the Excel workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = wb.active
        if sheet is None:
            raise HTTPException(status_code=400, detail="Workbook has no sheets")

        rows = []
        headers = None
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            if not any(row):
                continue
            rows.append([str(c).strip() if c is not None else "" for c in row])

        if not rows:
            return {"rows": 0, "headers": headers or [], "message": "No data rows found"}

        # Cache the rows for 7 days (cache_service auto-JSON-serializes)
        cache = get_cache_service()
        await cache.set(
            "inspection_forms:cached_rows",
            {"headers": headers, "rows": rows},
            ttl=7 * 24 * 3600,
        )

        logger.info(f"[INSPECTION-LETTERS] Cached {len(rows)} form rows from upload")
        return {
            "rows": len(rows),
            "headers": headers,
            "message": f"Cached {len(rows)} inspection responses",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[INSPECTION-LETTERS] Upload forms error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/inspection-letters/debug-forms")
async def debug_forms_sync(db: DbSession, current_user: CurrentUser):
    """Debug endpoint: show all drives, files, and permission checks."""
    from app.services.ms365_forms_sync_service import (
        MS365FormsSyncService,
        SP_SITE_HOST,
        SP_SITE_PATH,
        WORKBOOK_NAME,
        WORKBOOK_SHARING_URL,
        _encode_share_url,
    )

    debug = {
        "configured": MS365FormsSyncService.is_configured(),
        "steps": [],
    }
    if not debug["configured"]:
        debug["error"] = "MS365 not configured in env vars"
        return debug

    try:
        # Step 1: Site lookup
        try:
            site_data = await MS365FormsSyncService.graph_get(
                f"/sites/{SP_SITE_HOST}:{SP_SITE_PATH}"
            )
            debug["steps"].append({"step": "site_lookup", "ok": True, "site_id": site_data.get("id"), "name": site_data.get("displayName")})
            site_id = site_data["id"]
        except Exception as e:
            debug["steps"].append({"step": "site_lookup", "ok": False, "error": str(e)})
            return debug

        # Step 2: List all drives in the site
        try:
            drives_data = await MS365FormsSyncService.graph_get(f"/sites/{site_id}/drives")
            drives = drives_data.get("value", [])
            debug["steps"].append({
                "step": "list_drives",
                "ok": True,
                "count": len(drives),
                "drives": [{"id": d.get("id"), "name": d.get("name"), "driveType": d.get("driveType")} for d in drives],
            })
        except Exception as e:
            debug["steps"].append({"step": "list_drives", "ok": False, "error": str(e)})
            drives = []

        # Step 3: For each drive, try to find the workbook
        for drive in drives:
            drive_id = drive["id"]
            drive_name = drive.get("name", "?")
            try:
                item = await MS365FormsSyncService.graph_get(f"/drives/{drive_id}/root:/{WORKBOOK_NAME}")
                debug["steps"].append({"step": f"find_in_drive_{drive_name}", "ok": True, "item_id": item.get("id")})
            except Exception as e:
                debug["steps"].append({"step": f"find_in_drive_{drive_name}", "ok": False, "error": str(e)[:200]})

        # Step 4: Try listing children of each drive root
        for drive in drives[:3]:  # Limit to first 3 to avoid noise
            drive_id = drive["id"]
            drive_name = drive.get("name", "?")
            try:
                children = await MS365FormsSyncService.graph_get(f"/drives/{drive_id}/root/children")
                items = children.get("value", [])
                names = [c.get("name") for c in items]
                debug["steps"].append({
                    "step": f"list_children_{drive_name}",
                    "ok": True,
                    "count": len(items),
                    "names": names[:20],
                })
            except Exception as e:
                debug["steps"].append({"step": f"list_children_{drive_name}", "ok": False, "error": str(e)[:200]})

        # Step 5: Try /shares/ endpoint
        try:
            encoded = _encode_share_url(WORKBOOK_SHARING_URL)
            shares_data = await MS365FormsSyncService.graph_get(f"/shares/{encoded}/driveItem")
            debug["steps"].append({"step": "shares_endpoint", "ok": True, "item_id": shares_data.get("id"), "name": shares_data.get("name")})
        except Exception as e:
            debug["steps"].append({"step": "shares_endpoint", "ok": False, "error": str(e)[:300]})

        # Step 6: Try the configured drive ID directly
        from app.config import settings
        configured_drive = getattr(settings, "MS365_SHAREPOINT_DRIVE_ID", None)
        if configured_drive:
            try:
                drive_info = await MS365FormsSyncService.graph_get(f"/drives/{configured_drive}")
                debug["steps"].append({
                    "step": "configured_drive_info",
                    "ok": True,
                    "name": drive_info.get("name"),
                    "driveType": drive_info.get("driveType"),
                })
            except Exception as e:
                debug["steps"].append({"step": "configured_drive_info", "ok": False, "error": str(e)[:300]})

            # List children of configured drive root
            try:
                children = await MS365FormsSyncService.graph_get(f"/drives/{configured_drive}/root/children")
                items = children.get("value", [])
                names = [c.get("name") for c in items]
                debug["steps"].append({
                    "step": "configured_drive_root_children",
                    "ok": True,
                    "count": len(items),
                    "names": names[:30],
                })
            except Exception as e:
                debug["steps"].append({"step": "configured_drive_root_children", "ok": False, "error": str(e)[:300]})

            # Try to find the workbook in configured drive
            try:
                item = await MS365FormsSyncService.graph_get(f"/drives/{configured_drive}/root:/{WORKBOOK_NAME}")
                debug["steps"].append({"step": "find_in_configured_drive", "ok": True, "item_id": item.get("id")})
            except Exception as e:
                debug["steps"].append({"step": "find_in_configured_drive", "ok": False, "error": str(e)[:200]})

            # Search the configured drive
            try:
                import urllib.parse
                q = urllib.parse.quote("Septic")
                data = await MS365FormsSyncService.graph_get(f"/drives/{configured_drive}/root/search(q='{q}')")
                results = data.get("value", [])
                debug["steps"].append({
                    "step": "search_configured_drive",
                    "ok": True,
                    "count": len(results),
                    "matches": [{"name": r.get("name"), "id": r.get("id")} for r in results[:10]],
                })
            except Exception as e:
                debug["steps"].append({"step": "search_configured_drive", "ok": False, "error": str(e)[:300]})

        return debug
    except Exception as e:
        debug["fatal_error"] = str(e)
        return debug


class StandaloneLetterRequest(BaseModel):
    """Form data for generating a standalone inspection letter (no work order)."""
    customer_name: str
    customer_email: Optional[str] = None
    customer_phone: Optional[str] = None
    address: str
    inspection_date: str
    inspection_time: str = "12:00 PM CST"
    tank_location: str = ""
    tank_depth: str = ""
    permit_info: str = ""
    installation_year: str = ""
    system_type: str = "Standard septic system with one tank gravity flow"
    tank_capacity: str = "1000 gallons"
    flow_type: str = "gravity"
    condition: str = "good"
    pumped: bool = False
    baffles: str = ""
    operational_test: str = ""
    drain_field: str = "No signs of leaching up or super saturation. System appears to be functioning properly overall."
    notable_observations: str = ""
    pump_info: str = ""
    pump_chamber: str = ""
    homeowner_present: bool = False
    who_present: str = ""
    signer: str = "douglas_carter"


@router.post("/inspection-letters/sync-forms")
async def sync_inspection_forms_now(
    db: DbSession,
    current_user: CurrentUser,
):
    """Manually trigger sync of inspection form responses from SharePoint."""
    from app.services.ms365_forms_sync_service import MS365FormsSyncService
    if not MS365FormsSyncService.is_configured():
        raise HTTPException(status_code=503, detail="MS365 not configured")
    result = await MS365FormsSyncService.sync_inspection_forms()
    return result


@router.post("/inspection-letters/standalone/generate")
async def generate_standalone_letter(
    body: StandaloneLetterRequest,
    db: DbSession,
    current_user: CurrentUser,
):
    """Generate an AI inspection letter from manually entered form data (no work order)."""
    try:
        # Build a checklist dict that _build_inspection_text can consume
        checklist = {
            "address": body.address,
            "steps": {
                "1": {"address": body.address},
                "2": {
                    "tank_location": body.tank_location,
                    "tank_depth": body.tank_depth,
                },
                "3": {
                    "permit_info": body.permit_info,
                    "installation_year": body.installation_year,
                },
                "4": {
                    "system_type": f"{body.system_type}, {body.flow_type} flow",
                    "tank_capacity": body.tank_capacity,
                    "condition": body.condition,
                },
                "5": {
                    "pumped": "yes" if body.pumped else "no",
                    "baffles": body.baffles,
                    "operational_test": body.operational_test,
                },
                "6": {
                    "notable_observations": body.notable_observations,
                    "pump_info": body.pump_info,
                    "pump_chamber": body.pump_chamber,
                },
                "7": {
                    "drain_field": body.drain_field,
                },
            },
            "customFields": {
                "homeowner_present": body.homeowner_present,
            },
        }

        from app.services.inspection_letter_service import generate_letter_draft
        draft = await generate_letter_draft(checklist)

        return {
            "body": draft.get("body", ""),
            "generated_at": draft.get("generated_at"),
            "model": draft.get("model"),
            "status": draft.get("status", "draft"),
            "error": draft.get("error"),
            "form_data": {
                "customer_name": body.customer_name,
                "customer_email": body.customer_email,
                "customer_phone": body.customer_phone,
                "address": body.address,
                "inspection_date": body.inspection_date,
                "inspection_time": body.inspection_time,
                "signer": body.signer,
            },
        }
    except Exception as e:
        logger.error(f"[INSPECTION-LETTERS] Standalone generate error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/inspection-letters/standalone/pdf")
async def generate_standalone_letter_pdf(
    body: dict,
    db: DbSession,
    current_user: CurrentUser,
):
    """Render a standalone inspection letter as a PDF."""
    try:
        import base64
        from app.services.inspection_letter_service import render_letter_pdf

        letter_body = body.get("letter_body", "")
        if not letter_body:
            raise HTTPException(status_code=400, detail="letter_body is required")

        pdf_bytes = render_letter_pdf(
            letter_body=letter_body,
            customer_name=body.get("customer_name", "Valued Customer"),
            customer_address=body.get("address", ""),
            customer_email=body.get("customer_email", ""),
            customer_phone=body.get("customer_phone", ""),
            inspection_date=body.get("inspection_date", ""),
            inspection_time=body.get("inspection_time", "12:00 PM CST"),
            signer_key=body.get("signer", "douglas_carter"),
        )

        return {
            "pdf_base64": base64.b64encode(pdf_bytes).decode("ascii"),
            "status": "approved",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[INSPECTION-LETTERS] Standalone PDF error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/inspection-letters/{work_order_id}/generate")
async def generate_letter_for_wo(
    work_order_id: str,
    db: DbSession,
    current_user: CurrentUser,
):
    """Generate an AI letter draft for a specific work order (admin/office use)."""
    try:
        from sqlalchemy.orm import selectinload
        from app.services.inspection_letter_service import generate_letter_draft

        result = await db.execute(
            select(WorkOrder)
            .options(selectinload(WorkOrder.customer))
            .where(WorkOrder.id == work_order_id)
        )
        wo = result.scalars().first()
        if not wo:
            raise HTTPException(status_code=404, detail="Work order not found")

        checklist = wo.checklist if isinstance(wo.checklist, dict) else {}
        inspection = checklist.get("inspection") or {}

        # Build customer info for response
        customer_name = "Unknown"
        customer_email = None
        customer_phone = None
        if wo.customer:
            customer_name = f"{wo.customer.first_name or ''} {wo.customer.last_name or ''}".strip() or "Unknown"
            customer_email = wo.customer.email
            customer_phone = wo.customer.phone

        address = ""
        if wo.service_address_line1:
            parts = [wo.service_address_line1, wo.service_city, wo.service_state, wo.service_postal_code]
            address = ", ".join(p for p in parts if p)
        elif wo.customer:
            parts = [wo.customer.address_line1, wo.customer.city, wo.customer.state, wo.customer.postal_code]
            address = ", ".join(p for p in parts if p)

        insp_date = str(wo.scheduled_date) if wo.scheduled_date else ""
        insp_time = "12:00 PM CST"

        # Check if checklist has any usable inspection data
        has_steps = bool(inspection.get("steps")) or bool(checklist.get("steps"))

        # If no inspection data in checklist, try to fetch from MS Forms by customer/address
        forms_source = None
        if not has_steps:
            try:
                from app.services.ms365_forms_sync_service import MS365FormsSyncService
                if MS365FormsSyncService.is_configured():
                    logger.info(f"[INSPECTION-LETTERS] No checklist data for WO {wo.id}, trying Forms lookup")
                    form_data = await MS365FormsSyncService.find_inspection_by_address_or_name(
                        customer_name=customer_name,
                        address=address,
                    )
                    if form_data:
                        from app.services.inspection_letter_service import form_data_to_checklist
                        checklist = form_data_to_checklist(form_data)
                        forms_source = "ms_forms"
                        # Also save the form data to the work order for future use
                        wo_checklist = wo.checklist if isinstance(wo.checklist, dict) else {}
                        wo_checklist["inspection"] = checklist
                        wo_checklist["inspection"]["source"] = "ms_forms"
                        wo.checklist = wo_checklist
                        from sqlalchemy.orm.attributes import flag_modified
                        flag_modified(wo, "checklist")
                        await db.commit()
                        logger.info(f"[INSPECTION-LETTERS] Backfilled WO {wo.id} with MS Forms data")
            except Exception as forms_err:
                logger.warning(f"[INSPECTION-LETTERS] MS Forms fallback failed: {forms_err}")

        # generate_letter_draft expects the inspection dict (with steps at top
        # level), not the outer work-order checklist. Pull out inspection and
        # fall back to checklist itself if the data happens to live there.
        letter_input = checklist.get("inspection") if isinstance(checklist.get("inspection"), dict) else None
        if not letter_input or not letter_input.get("steps"):
            letter_input = checklist if checklist.get("steps") else (letter_input or {})
        if not letter_input.get("address"):
            letter_input["address"] = address
        if not letter_input.get("customer_name"):
            letter_input["customer_name"] = customer_name

        draft = await generate_letter_draft(letter_input)

        # Store draft in checklist
        if draft.get("body"):
            inspection["ai_letter"] = draft
            checklist["inspection"] = inspection
            wo.checklist = checklist
            from sqlalchemy.orm.attributes import flag_modified
            flag_modified(wo, "checklist")
            await db.commit()

        return {
            "body": draft.get("body", ""),
            "generated_at": draft.get("generated_at"),
            "model": draft.get("model"),
            "status": draft.get("status", "draft"),
            "error": draft.get("error"),
            "source": forms_source or "checklist",
            "form_data": {
                "customer_name": customer_name,
                "customer_email": customer_email or "",
                "customer_phone": customer_phone or "",
                "address": address,
                "inspection_date": insp_date,
                "inspection_time": insp_time,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[INSPECTION-LETTERS] WO generate error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/inspection-letters/{work_order_id}/pdf")
async def generate_letter_pdf_for_wo(
    work_order_id: str,
    body: dict,
    db: DbSession,
    current_user: CurrentUser,
):
    """Render letter PDF for a work order, including photos."""
    try:
        import base64
        from sqlalchemy.orm import selectinload
        from app.services.inspection_letter_service import render_letter_pdf
        from app.models.work_order_photo import WorkOrderPhoto

        letter_body = body.get("letter_body", "")
        if not letter_body:
            raise HTTPException(status_code=400, detail="letter_body is required")

        result = await db.execute(
            select(WorkOrder)
            .options(selectinload(WorkOrder.customer))
            .where(WorkOrder.id == work_order_id)
        )
        wo = result.scalars().first()
        if not wo:
            raise HTTPException(status_code=404, detail="Work order not found")

        customer_name = body.get("customer_name", "Valued Customer")
        customer_address = body.get("address", "")
        if not customer_address and wo.customer:
            parts = [wo.customer.address_line1, wo.customer.city, wo.customer.state, wo.customer.postal_code]
            customer_address = ", ".join(p for p in parts if p)

        # Fetch photos
        photo_result = await db.execute(
            select(WorkOrderPhoto)
            .where(WorkOrderPhoto.work_order_id == work_order_id)
            .order_by(WorkOrderPhoto.created_at)
        )
        photo_rows = photo_result.scalars().all()
        photos = [{"data": p.data, "photo_type": p.photo_type} for p in photo_rows if p.data]

        pdf_bytes = render_letter_pdf(
            letter_body=letter_body,
            customer_name=customer_name,
            customer_address=customer_address,
            customer_email=body.get("customer_email", ""),
            customer_phone=body.get("customer_phone", ""),
            inspection_date=body.get("inspection_date", ""),
            inspection_time=body.get("inspection_time", "12:00 PM CST"),
            signer_key=body.get("signer", "douglas_carter"),
            photos=photos,
        )

        # Store in checklist
        checklist = wo.checklist if isinstance(wo.checklist, dict) else {}
        inspection = checklist.get("inspection") or {}
        inspection.setdefault("ai_letter", {})
        inspection["ai_letter"]["status"] = "approved"
        inspection["ai_letter"]["approved_body"] = letter_body
        inspection["ai_letter"]["signer"] = body.get("signer", "douglas_carter")
        checklist["inspection"] = inspection
        wo.checklist = checklist
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(wo, "checklist")
        await db.commit()

        return {
            "pdf_base64": base64.b64encode(pdf_bytes).decode("ascii"),
            "status": "approved",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[INSPECTION-LETTERS] WO PDF error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/fix-billing-column")
async def fix_billing_customer_column(db: DbSession, current_user: CurrentUser):
    """Add billing_customer_id column if missing (migration 093 safety net)."""
    try:
        result = await db.execute(
            text("SELECT column_name FROM information_schema.columns WHERE table_name = 'work_orders' AND column_name = 'billing_customer_id'")
        )
        if not result.fetchone():
            await db.execute(text("ALTER TABLE work_orders ADD COLUMN billing_customer_id UUID REFERENCES customers(id) ON DELETE SET NULL"))
            await db.execute(text("CREATE INDEX IF NOT EXISTS ix_work_orders_billing_customer_id ON work_orders(billing_customer_id)"))
            await db.commit()
            return {"status": "success", "message": "billing_customer_id column added"}
        return {"status": "ok", "message": "billing_customer_id column already exists"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@router.post("/fix-table")
async def fix_work_orders_table(db: DbSession, current_user: CurrentUser):
    """Add work_order_number column and backfill existing work orders."""
    try:
        # Check if column exists
        result = await db.execute(
            text(
                """SELECT column_name FROM information_schema.columns
                WHERE table_name = 'work_orders' AND column_name = 'work_order_number'"""
            )
        )
        exists = result.fetchone()

        if not exists:
            logger.info("Adding work_order_number column...")
            await db.execute(
                text("ALTER TABLE work_orders ADD COLUMN work_order_number VARCHAR(20)")
            )
            await db.commit()

            # Backfill existing work orders
            logger.info("Backfilling work order numbers...")
            await db.execute(
                text("""
                    WITH numbered AS (
                        SELECT id, ROW_NUMBER() OVER (ORDER BY created_at NULLS LAST, id) as rn
                        FROM work_orders
                        WHERE work_order_number IS NULL
                    )
                    UPDATE work_orders wo
                    SET work_order_number = 'WO-' || LPAD(n.rn::text, 6, '0')
                    FROM numbered n
                    WHERE wo.id = n.id
                """)
            )
            await db.commit()

            # Add index
            try:
                await db.execute(
                    text("CREATE UNIQUE INDEX IF NOT EXISTS ix_work_orders_number ON work_orders(work_order_number)")
                )
                await db.commit()
            except Exception:
                pass

            return {"status": "success", "message": "work_order_number column added and backfilled"}
        else:
            # Check for any NULL values and backfill them
            result = await db.execute(
                text("SELECT COUNT(*) FROM work_orders WHERE work_order_number IS NULL")
            )
            null_count = result.scalar()

            if null_count and null_count > 0:
                logger.info(f"Backfilling {null_count} work orders with NULL work_order_number...")
                await db.execute(
                    text("""
                        WITH max_num AS (
                            SELECT COALESCE(MAX(CAST(REPLACE(work_order_number, 'WO-', '') AS INTEGER)), 0) as max_n
                            FROM work_orders
                            WHERE work_order_number IS NOT NULL
                        ),
                        numbered AS (
                            SELECT id, ROW_NUMBER() OVER (ORDER BY created_at NULLS LAST, id) as rn
                            FROM work_orders
                            WHERE work_order_number IS NULL
                        )
                        UPDATE work_orders wo
                        SET work_order_number = 'WO-' || LPAD((n.rn + (SELECT max_n FROM max_num))::text, 6, '0')
                        FROM numbered n
                        WHERE wo.id = n.id
                    """)
                )
                await db.commit()
                return {"status": "success", "message": f"Backfilled {null_count} work orders"}

            return {"status": "success", "message": "Column already exists, no action needed"}

    except Exception as e:
        logger.error(f"Error fixing work_orders table: {e}")
        return {"status": "error", "message": str(e)}


def _customer_embed(c: Customer) -> dict:
    """Build a CustomerEmbed dict from a Customer model."""
    return {
        "id": str(c.id),
        "first_name": c.first_name or "",
        "last_name": c.last_name or "",
        "email": c.email,
        "phone": c.phone,
        "address_line1": c.address_line1,
        "city": c.city,
        "state": c.state,
        "postal_code": c.postal_code,
    }


def work_order_with_customer_name(wo: WorkOrder, customer: Optional[Customer], billing_customer: Optional[Customer] = None) -> dict:
    """Convert WorkOrder to dict with customer_name populated from Customer JOIN."""
    customer_name = None
    customer_phone = None
    customer_email = None
    customer_obj = None
    if customer:
        first = customer.first_name or ""
        last = customer.last_name or ""
        customer_name = f"{first} {last}".strip() or None
        customer_phone = customer.phone or None
        customer_email = customer.email or None
        customer_obj = _customer_embed(customer)

    # Build billing customer embed
    billing_customer_obj = None
    if billing_customer:
        billing_customer_obj = _customer_embed(billing_customer)

    return {
        "id": wo.id,
        "work_order_number": wo.work_order_number,
        "customer_id": wo.customer_id,
        "billing_customer_id": wo.billing_customer_id,
        "customer_name": customer_name,
        "customer_phone": customer_phone,
        "customer_email": customer_email,
        "customer": customer_obj,
        "billing_customer": billing_customer_obj,
        "technician_id": wo.technician_id,
        "job_type": str(wo.job_type) if wo.job_type else None,
        "status": str(wo.status) if wo.status else "draft",
        "priority": str(wo.priority) if wo.priority else "normal",
        "scheduled_date": wo.scheduled_date,
        "time_window_start": wo.time_window_start,
        "time_window_end": wo.time_window_end,
        "estimated_duration_hours": wo.estimated_duration_hours or {
            "inspection": 0.5, "pumping": 1.0, "repair": 2.0,
            "installation": 4.0, "maintenance": 1.0, "grease_trap": 1.0, "emergency": 2.0,
        }.get(str(wo.job_type) if wo.job_type else "", 1.0),
        "service_address_line1": wo.service_address_line1,
        "service_address_line2": wo.service_address_line2,
        "service_city": wo.service_city,
        "service_state": wo.service_state,
        "service_postal_code": wo.service_postal_code,
        "service_latitude": wo.service_latitude,
        "service_longitude": wo.service_longitude,
        "estimated_gallons": wo.estimated_gallons,
        "notes": wo.notes,
        "internal_notes": wo.internal_notes,
        "is_recurring": wo.is_recurring,
        "recurrence_frequency": wo.recurrence_frequency,
        "next_recurrence_date": wo.next_recurrence_date,
        "checklist": wo.checklist,
        "assigned_vehicle": wo.assigned_vehicle,
        "assigned_technician": wo.assigned_technician,
        "system_type": wo.system_type or "conventional",
        "total_amount": wo.total_amount,
        "created_at": wo.created_at,
        "updated_at": wo.updated_at,
        "actual_start_time": wo.actual_start_time,
        "actual_end_time": wo.actual_end_time,
        "travel_start_time": wo.travel_start_time,
        "travel_end_time": wo.travel_end_time,
        "break_minutes": wo.break_minutes,
        "total_labor_minutes": wo.total_labor_minutes,
        "total_travel_minutes": wo.total_travel_minutes,
        "is_clocked_in": wo.is_clocked_in,
        "clock_in_gps_lat": wo.clock_in_gps_lat,
        "clock_in_gps_lon": wo.clock_in_gps_lon,
        "clock_out_gps_lat": wo.clock_out_gps_lat,
        "clock_out_gps_lon": wo.clock_out_gps_lon,
    }


@router.get("", response_model=WorkOrderListResponse)
async def list_work_orders(
    db: DbSession,
    current_user: CurrentUser,
    entity: EntityCtx,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=1000),
    customer_id: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    job_type: Optional[str] = None,
    priority: Optional[str] = None,
    assigned_technician: Optional[str] = None,
    technician_id: Optional[str] = None,
    scheduled_date: Optional[str] = None,
    scheduled_date_from: Optional[datetime] = None,
    scheduled_date_to: Optional[datetime] = None,
):
    """List work orders with pagination, filtering, and real customer names."""
    # Check cache first
    cache = get_cache_service()
    cache_key = f"workorders:list:{page}:{page_size}:{customer_id}:{status_filter}:{job_type}:{priority}:{assigned_technician}:{technician_id}:{scheduled_date}:{scheduled_date_from}:{scheduled_date_to}"
    cached = await cache.get(cache_key)
    if cached is not None:
        return cached

    try:
        # Base query with LEFT JOIN to Customer + BillingCustomer
        query = (
            select(WorkOrder, Customer, BillingCustomer)
            .outerjoin(Customer, WorkOrder.customer_id == Customer.id)
            .outerjoin(BillingCustomer, WorkOrder.billing_customer_id == BillingCustomer.id)
        )

        # Apply filters
        if customer_id:
            query = query.where(WorkOrder.customer_id == customer_id)
        if status_filter:
            # Cast status column to string for comparison (handles PostgreSQL ENUM)
            query = query.where(cast(WorkOrder.status, String) == status_filter)
        if job_type:
            query = query.where(cast(WorkOrder.job_type, String) == job_type)
        if priority:
            query = query.where(cast(WorkOrder.priority, String) == priority)
        if assigned_technician:
            query = query.where(WorkOrder.assigned_technician == assigned_technician)
        if technician_id:
            query = query.where(WorkOrder.technician_id == technician_id)
        if scheduled_date:
            # Parse string to date object for proper comparison
            try:
                date_obj = date_type.fromisoformat(scheduled_date)
                query = query.where(WorkOrder.scheduled_date == date_obj)
            except ValueError:
                pass  # Invalid date format, skip filter
        if scheduled_date_from:
            query = query.where(WorkOrder.scheduled_date >= scheduled_date_from)
        if scheduled_date_to:
            query = query.where(WorkOrder.scheduled_date <= scheduled_date_to)

        # Multi-entity filtering
        if entity:
            if entity.is_default:
                query = query.where(or_(WorkOrder.entity_id == entity.id, WorkOrder.entity_id == None))
            else:
                query = query.where(WorkOrder.entity_id == entity.id)

        # Get total count - simple count with same filters
        count_query = select(func.count()).select_from(WorkOrder)
        if customer_id:
            count_query = count_query.where(WorkOrder.customer_id == customer_id)
        if status_filter:
            count_query = count_query.where(cast(WorkOrder.status, String) == status_filter)
        if job_type:
            count_query = count_query.where(cast(WorkOrder.job_type, String) == job_type)
        if priority:
            count_query = count_query.where(cast(WorkOrder.priority, String) == priority)
        if assigned_technician:
            count_query = count_query.where(WorkOrder.assigned_technician == assigned_technician)
        if technician_id:
            count_query = count_query.where(WorkOrder.technician_id == technician_id)
        if scheduled_date:
            try:
                date_obj = date_type.fromisoformat(scheduled_date)
                count_query = count_query.where(WorkOrder.scheduled_date == date_obj)
            except ValueError:
                pass
        if scheduled_date_from:
            count_query = count_query.where(WorkOrder.scheduled_date >= scheduled_date_from)
        if scheduled_date_to:
            count_query = count_query.where(WorkOrder.scheduled_date <= scheduled_date_to)
        if entity:
            if entity.is_default:
                count_query = count_query.where(or_(WorkOrder.entity_id == entity.id, WorkOrder.entity_id == None))
            else:
                count_query = count_query.where(WorkOrder.entity_id == entity.id)

        total_result = await db.execute(count_query)
        total = total_result.scalar() or 0

        # Apply pagination
        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size).order_by(WorkOrder.created_at.desc())

        # Execute query - returns tuples of (WorkOrder, Customer)
        result = await db.execute(query)
        rows = result.all()

        # Convert to dicts with customer_name populated
        items = [work_order_with_customer_name(wo, customer, billing_cust) for wo, customer, billing_cust in rows]

        response = WorkOrderListResponse(
            items=items,
            total=total,
            page=page,
            page_size=page_size,
        )
        await cache.set(cache_key, jsonable_encoder(response), ttl=TTL.SHORT)
        return response
    except Exception as e:
        logger.error(f"Error in list_work_orders: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


@router.get("/cursor", response_model=WorkOrderCursorResponse)
async def list_work_orders_cursor(
    db: DbSession,
    current_user: CurrentUser,
    cursor: Optional[str] = None,
    page_size: int = Query(20, ge=1, le=100),
    customer_id: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    job_type: Optional[str] = None,
    priority: Optional[str] = None,
    assigned_technician: Optional[str] = None,
    technician_id: Optional[str] = None,
    scheduled_date: Optional[str] = None,
    scheduled_date_from: Optional[datetime] = None,
    scheduled_date_to: Optional[datetime] = None,
):
    """List work orders with cursor-based pagination (efficient for large datasets).

    Uses cursor pagination instead of offset pagination for better performance
    when paginating through large result sets.
    """
    try:
        # Base query with LEFT JOIN to Customer + BillingCustomer
        query = (
            select(WorkOrder, Customer, BillingCustomer)
            .outerjoin(Customer, WorkOrder.customer_id == Customer.id)
            .outerjoin(BillingCustomer, WorkOrder.billing_customer_id == BillingCustomer.id)
        )

        # Apply filters
        if customer_id:
            query = query.where(WorkOrder.customer_id == customer_id)
        if status_filter:
            query = query.where(cast(WorkOrder.status, String) == status_filter)
        if job_type:
            query = query.where(cast(WorkOrder.job_type, String) == job_type)
        if priority:
            query = query.where(cast(WorkOrder.priority, String) == priority)
        if assigned_technician:
            query = query.where(WorkOrder.assigned_technician == assigned_technician)
        if technician_id:
            query = query.where(WorkOrder.technician_id == technician_id)
        if scheduled_date:
            try:
                date_obj = date_type.fromisoformat(scheduled_date)
                query = query.where(WorkOrder.scheduled_date == date_obj)
            except ValueError:
                pass
        if scheduled_date_from:
            query = query.where(WorkOrder.scheduled_date >= scheduled_date_from)
        if scheduled_date_to:
            query = query.where(WorkOrder.scheduled_date <= scheduled_date_to)

        # Apply cursor filter if provided
        if cursor:
            try:
                cursor_id, cursor_ts = decode_cursor(cursor)
                # Descending order: get items BEFORE the cursor
                if cursor_ts:
                    cursor_filter = or_(
                        WorkOrder.created_at < cursor_ts,
                        and_(WorkOrder.created_at == cursor_ts, WorkOrder.id < cursor_id),
                    )
                else:
                    cursor_filter = WorkOrder.id < cursor_id
                query = query.where(cursor_filter)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid cursor format")

        # Order by created_at descending, then id for stable pagination
        query = query.order_by(WorkOrder.created_at.desc(), WorkOrder.id.desc())

        # Fetch one extra to determine if there are more results
        query = query.limit(page_size + 1)

        result = await db.execute(query)
        rows = result.all()

        # Check if there are more results
        has_more = len(rows) > page_size
        rows = rows[:page_size]  # Trim to requested page size

        # Convert to dicts with customer_name populated
        items = [work_order_with_customer_name(wo, customer, billing_cust) for wo, customer, billing_cust in rows]

        # Build next cursor from last item
        next_cursor = None
        if has_more and rows:
            last_wo, _, _ = rows[-1]
            next_cursor = encode_cursor(last_wo.id, last_wo.created_at)

        return WorkOrderCursorResponse(
            items=items,
            next_cursor=next_cursor,
            has_more=has_more,
            total=None,  # Omit total for cursor pagination (expensive to compute)
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in list_work_orders_cursor: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


# ============================================
# Bulk Operations
# ============================================

VALID_STATUSES = {"draft", "scheduled", "confirmed", "enroute", "on_site", "in_progress", "completed", "canceled", "requires_followup"}
MAX_BULK_SIZE = 200


class BulkStatusRequest(BaseModel):
    """Bulk update status for multiple work orders."""
    ids: List[str] = Field(..., max_length=MAX_BULK_SIZE)
    status: str


class BulkAssignRequest(BaseModel):
    """Bulk assign technician to multiple work orders."""
    ids: List[str] = Field(..., max_length=MAX_BULK_SIZE)
    assigned_technician: Optional[str] = None
    technician_id: Optional[str] = None


class BulkDeleteRequest(BaseModel):
    """Bulk delete multiple work orders."""
    ids: List[str] = Field(..., max_length=MAX_BULK_SIZE)


class BulkResult(BaseModel):
    """Result of a bulk operation."""
    success_count: int
    failed_count: int
    errors: List[dict] = []


@router.patch("/bulk/status")
async def bulk_update_status(
    request: BulkStatusRequest,
    db: DbSession,
    current_user: CurrentUser,
) -> BulkResult:
    """Bulk update work order status. Max 200 at a time."""
    if request.status not in VALID_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status: {request.status}")

    success = 0
    errors = []

    for wo_id in request.ids:
        try:
            result = await db.execute(select(WorkOrder).where(WorkOrder.id == wo_id))
            wo = result.scalar_one_or_none()
            if not wo:
                errors.append({"id": wo_id, "error": "Not found"})
                continue
            wo.status = request.status
            wo.updated_at = datetime.utcnow()
            success += 1
        except Exception as e:
            errors.append({"id": wo_id, "error": str(e)})

    await db.commit()

    # Invalidate cache
    cache = get_cache_service()
    await cache.delete_pattern("work-orders:*")

    # Broadcast WebSocket event
    await manager.broadcast({
        "type": "work_order_update",
        "data": {"count": success, "status": request.status},
    })

    return BulkResult(success_count=success, failed_count=len(errors), errors=errors)


@router.patch("/bulk/assign")
async def bulk_assign_technician(
    request: BulkAssignRequest,
    db: DbSession,
    current_user: CurrentUser,
) -> BulkResult:
    """Bulk assign a technician to multiple work orders. Max 200."""
    # Resolve technician_id from name if not provided
    tech_id = None
    if request.technician_id:
        tech_id = request.technician_id
    elif request.assigned_technician:
        parts = request.assigned_technician.strip().split()
        if len(parts) >= 2:
            tech_result = await db.execute(
                select(Technician).where(
                    func.lower(Technician.first_name) == parts[0].lower(),
                    func.lower(Technician.last_name) == parts[-1].lower(),
                )
            )
            tech = tech_result.scalar_one_or_none()
            if tech:
                tech_id = str(tech.id)

    success = 0
    errors = []

    for wo_id in request.ids:
        try:
            result = await db.execute(select(WorkOrder).where(WorkOrder.id == wo_id))
            wo = result.scalar_one_or_none()
            if not wo:
                errors.append({"id": wo_id, "error": "Not found"})
                continue
            wo.assigned_technician = request.assigned_technician
            if tech_id:
                wo.technician_id = tech_id
            wo.updated_at = datetime.utcnow()
            success += 1
        except Exception as e:
            errors.append({"id": wo_id, "error": str(e)})

    await db.commit()

    cache = get_cache_service()
    await cache.delete_pattern("work-orders:*")

    await manager.broadcast({
        "type": "dispatch_update",
        "data": {"count": success, "technician": request.assigned_technician},
    })

    return BulkResult(success_count=success, failed_count=len(errors), errors=errors)


@router.delete("/bulk")
async def bulk_delete_work_orders(
    request: BulkDeleteRequest,
    db: DbSession,
    current_user: CurrentUser,
) -> BulkResult:
    """Bulk delete work orders. Max 200."""
    success = 0
    errors = []

    for wo_id in request.ids:
        try:
            result = await db.execute(select(WorkOrder).where(WorkOrder.id == wo_id))
            wo = result.scalar_one_or_none()
            if not wo:
                errors.append({"id": wo_id, "error": "Not found"})
                continue
            await db.delete(wo)
            success += 1
        except Exception as e:
            errors.append({"id": wo_id, "error": str(e)})

    await db.commit()

    cache = get_cache_service()
    await cache.delete_pattern("work-orders:*")

    return BulkResult(success_count=success, failed_count=len(errors), errors=errors)


# ============================================
# Route Optimization
# ============================================

import math


def _haversine_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate great-circle distance between two points in miles."""
    R = 3958.8  # Earth radius in miles
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(dlon / 2) ** 2
    )
    return R * 2 * math.asin(math.sqrt(a))


def _nearest_neighbor_route(
    jobs: list[dict], start_lat: float, start_lng: float
) -> tuple[list[dict], float]:
    """Nearest-neighbor TSP that respects time_window_start constraints.

    Jobs WITH a time_window_start are pinned at their correct chronological
    position. Jobs WITHOUT a time window are slotted into the remaining gaps
    using nearest-neighbor distance optimization.

    This means: if Tod has a 9:00 AM job and a 1:00 PM job, the optimizer
    routes flexible jobs between those two fixed appointments.
    """
    from datetime import time as time_type

    def _parse_time(t) -> time_type | None:
        """Parse a time_window_start value (time obj, str, or None)."""
        if t is None:
            return None
        if isinstance(t, time_type):
            return t
        if isinstance(t, str):
            try:
                parts = t.split(":")
                return time_type(int(parts[0]), int(parts[1]))
            except (ValueError, IndexError):
                return None
        return None

    # Separate time-pinned vs flexible jobs
    pinned = []
    flexible = []
    for j in jobs:
        tw = _parse_time(j.get("time_window_start"))
        if tw is not None:
            pinned.append((tw, j))
        else:
            flexible.append(j)

    # Sort pinned jobs by time
    pinned.sort(key=lambda x: x[0])

    # Build final route: interleave pinned jobs with nearest-neighbor flexible fills
    ordered = []
    total_dist = 0.0
    current_lat, current_lng = start_lat, start_lng

    for _tw, pinned_job in pinned:
        # Before each pinned job, greedily fill with nearest flexible jobs
        # that are closer to current position than the pinned job
        pinned_dist = _haversine_miles(current_lat, current_lng, pinned_job["lat"], pinned_job["lng"])
        while flexible:
            nearest_flex = min(
                flexible,
                key=lambda j: _haversine_miles(current_lat, current_lng, j["lat"], j["lng"]),
            )
            flex_dist = _haversine_miles(current_lat, current_lng, nearest_flex["lat"], nearest_flex["lng"])
            # Only insert flexible job if it's on the way (closer than pinned)
            if flex_dist < pinned_dist * 0.6:
                total_dist += flex_dist
                current_lat, current_lng = nearest_flex["lat"], nearest_flex["lng"]
                ordered.append(nearest_flex)
                flexible.remove(nearest_flex)
                # Recalculate pinned distance from new position
                pinned_dist = _haversine_miles(current_lat, current_lng, pinned_job["lat"], pinned_job["lng"])
            else:
                break

        # Add the pinned job
        dist = _haversine_miles(current_lat, current_lng, pinned_job["lat"], pinned_job["lng"])
        total_dist += dist
        current_lat, current_lng = pinned_job["lat"], pinned_job["lng"]
        ordered.append(pinned_job)

    # Append remaining flexible jobs via nearest-neighbor
    while flexible:
        nearest = min(
            flexible,
            key=lambda j: _haversine_miles(current_lat, current_lng, j["lat"], j["lng"]),
        )
        dist = _haversine_miles(current_lat, current_lng, nearest["lat"], nearest["lng"])
        total_dist += dist
        current_lat, current_lng = nearest["lat"], nearest["lng"]
        ordered.append(nearest)
        flexible.remove(nearest)

    return ordered, total_dist


def _address_to_approx_coords(address: str) -> tuple[float, float]:
    """
    Deterministic address-based approximation for San Marcos TX area.
    Used when no stored coordinates are available.
    """
    h = hash(address) % 10000
    lat = 29.8 + (h % 100) / 500  # ~29.8 to 30.0
    lng = -97.9 + (h // 100 % 100) / 500  # ~-97.9 to -97.7
    return lat, lng


class RouteOptimizeRequest(BaseModel):
    job_ids: list[str]
    start_lat: Optional[float] = None
    start_lng: Optional[float] = None
    start_address: Optional[str] = "105 S Comanche St, San Marcos, TX 78666"


class RouteOptimizeResponse(BaseModel):
    ordered_job_ids: list[str]
    total_distance_miles: float
    estimated_drive_minutes: int
    waypoints: list[dict]  # [{job_id, address, lat, lng}]


@router.post("/optimize-route", response_model=RouteOptimizeResponse)
async def optimize_route(
    request: RouteOptimizeRequest,
    db: DbSession,
    current_user: CurrentUser,
):
    """
    Given a list of job IDs and a start location, return the jobs
    reordered by nearest-neighbor haversine distance.

    Input: { "job_ids": [...], "start_lat": 30.0, "start_lng": -97.0 }
    OR: { "job_ids": [...], "start_address": "123 Main St, San Marcos TX" }

    Output: {
        "ordered_job_ids": [...],
        "total_distance_miles": 47.3,
        "estimated_drive_minutes": 68,
        "waypoints": [{"job_id": ..., "address": ..., "lat": ..., "lng": ...}]
    }
    """
    if not request.job_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="job_ids must not be empty",
        )

    # Determine start coordinates
    if request.start_lat is not None and request.start_lng is not None:
        start_lat = request.start_lat
        start_lng = request.start_lng
    else:
        addr = request.start_address or "105 S Comanche St, San Marcos, TX 78666"
        start_lat, start_lng = _address_to_approx_coords(addr)

    # Fetch work orders with customer data for coordinates
    query = (
        select(WorkOrder, Customer)
        .outerjoin(Customer, WorkOrder.customer_id == Customer.id)
        .where(WorkOrder.id.in_(request.job_ids))
    )
    result = await db.execute(query)
    rows = result.all()

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No work orders found for the provided job_ids",
        )

    # Build job dicts with coordinates
    jobs = []
    for wo, customer in rows:
        # Prefer work order service coords, then customer coords, then hash approximation
        lat = None
        lng = None

        if wo.service_latitude is not None and wo.service_longitude is not None:
            lat = float(wo.service_latitude)
            lng = float(wo.service_longitude)
        elif customer and customer.latitude is not None and customer.longitude is not None:
            lat = float(customer.latitude)
            lng = float(customer.longitude)
        else:
            # Build address string for hash approximation
            address_parts = [
                wo.service_address_line1 or (customer.address_line1 if customer else None),
                wo.service_city or (customer.city if customer else None),
                wo.service_state or (customer.state if customer else None),
                wo.service_postal_code or (customer.postal_code if customer else None),
            ]
            address_str = ", ".join(p for p in address_parts if p)
            lat, lng = _address_to_approx_coords(address_str or str(wo.id))

        # Build human-readable address
        addr_parts = [
            wo.service_address_line1,
            wo.service_city,
            wo.service_state,
            wo.service_postal_code,
        ]
        address = ", ".join(p for p in addr_parts if p) or "Unknown address"

        jobs.append({
            "job_id": str(wo.id),
            "address": address,
            "lat": lat,
            "lng": lng,
            "time_window_start": wo.time_window_start,
        })

    # Run nearest-neighbor optimization
    ordered_jobs, total_distance = _nearest_neighbor_route(jobs, start_lat, start_lng)

    # Estimate drive time: assume average 35 mph for rural/suburban Texas
    estimated_drive_minutes = int(round(total_distance / 35 * 60))

    ordered_job_ids = [j["job_id"] for j in ordered_jobs]
    waypoints = [
        {
            "job_id": j["job_id"],
            "address": j["address"],
            "lat": j["lat"],
            "lng": j["lng"],
        }
        for j in ordered_jobs
    ]

    return RouteOptimizeResponse(
        ordered_job_ids=ordered_job_ids,
        total_distance_miles=round(total_distance, 2),
        estimated_drive_minutes=estimated_drive_minutes,
        waypoints=waypoints,
    )


# ============================================
# Single Work Order Operations
# ============================================


@router.get("/{work_order_id}", response_model=WorkOrderResponse)
async def get_work_order(
    work_order_id: str,
    db: DbSession,
    current_user: CurrentUser,
):
    """Get a single work order by ID with customer name."""
    # JOIN with Customer + BillingCustomer
    query = (
        select(WorkOrder, Customer, BillingCustomer)
        .outerjoin(Customer, WorkOrder.customer_id == Customer.id)
        .outerjoin(BillingCustomer, WorkOrder.billing_customer_id == BillingCustomer.id)
        .where(WorkOrder.id == work_order_id)
    )
    result = await db.execute(query)
    row = result.first()

    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Work order not found",
        )

    work_order, customer, billing_cust = row
    return work_order_with_customer_name(work_order, customer, billing_cust)


async def generate_work_order_number(db: DbSession) -> str:
    """Generate next work order number in WO-NNNNNN format."""
    result = await db.execute(
        select(func.max(WorkOrder.work_order_number))
    )
    last_number = result.scalar()
    if last_number and last_number.startswith("WO-"):
        try:
            num = int(last_number.replace("WO-", "")) + 1
        except ValueError:
            num = 1
    else:
        num = 1
    return f"WO-{num:06d}"


@router.post("", response_model=WorkOrderResponse, status_code=status.HTTP_201_CREATED)
async def create_work_order(
    work_order_data: WorkOrderCreate,
    request: Request,
    db: DbSession,
    current_user: CurrentUser,
    entity: EntityCtx,
):
    """Create a new work order."""
    data = work_order_data.model_dump()
    data["id"] = str(uuid.uuid4())
    if entity:
        data["entity_id"] = str(entity.id)
    data["work_order_number"] = await generate_work_order_number(db)

    # Audit trail fields
    data["created_by"] = current_user.email if current_user else None
    data["updated_by"] = current_user.email if current_user else None
    data["source"] = request.headers.get("X-Source", "crm")
    data["created_at"] = datetime.utcnow()
    data["updated_at"] = datetime.utcnow()

    # Set default estimated_duration_hours based on job type if not provided
    if not data.get("estimated_duration_hours"):
        job_type_durations = {
            "inspection": 0.5,   # 30 minutes
            "pumping": 1.0,      # 1 hour
            "repair": 2.0,       # 2 hours
            "installation": 4.0, # 4 hours
            "maintenance": 1.0,  # 1 hour
            "grease_trap": 1.0,  # 1 hour
            "emergency": 2.0,    # 2 hours
        }
        data["estimated_duration_hours"] = job_type_durations.get(
            data.get("job_type", ""), 1.0
        )

    # Auto-resolve assigned_technician → technician_id if not already set
    if data.get("assigned_technician") and not data.get("technician_id"):
        tech_name = data["assigned_technician"].strip()
        name_parts = tech_name.split(None, 1)
        if len(name_parts) >= 2:
            tech_result = await db.execute(
                select(Technician).where(
                    Technician.first_name == name_parts[0],
                    Technician.last_name == name_parts[1],
                    Technician.is_active == True,
                ).limit(1)
            )
        else:
            tech_result = await db.execute(
                select(Technician).where(
                    Technician.first_name == tech_name,
                    Technician.is_active == True,
                ).limit(1)
            )
        matched_tech = tech_result.scalar_one_or_none()
        if matched_tech:
            data["technician_id"] = str(matched_tech.id)
            logger.info(f"Auto-resolved technician '{tech_name}' → {matched_tech.id}")

    work_order = WorkOrder(**data)
    db.add(work_order)
    await db.commit()
    await db.refresh(work_order)

    # Create audit log entry
    try:
        client_ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else None)
        audit_entry = WorkOrderAuditLog(
            work_order_id=work_order.id,
            action="created",
            description=f"Work order {work_order.work_order_number} created for {work_order.job_type}",
            user_email=current_user.email if current_user else None,
            user_name=getattr(current_user, "full_name", None) or (current_user.email if current_user else "System"),
            source=request.headers.get("X-Source", "crm"),
            ip_address=client_ip,
            user_agent=request.headers.get("User-Agent", None),
            changes={"status": {"old": None, "new": work_order.status}, "job_type": {"old": None, "new": work_order.job_type}},
        )
        db.add(audit_entry)
        await db.commit()
    except Exception as e:
        logger.warning(f"Failed to create audit log for WO {work_order.id}: {e}")

    # Broadcast work order created event via WebSocket
    await manager.broadcast_event(
        event_type="work_order_update",
        data={
            "id": work_order.id,
            "customer_id": str(work_order.customer_id),
            "job_type": str(work_order.job_type) if work_order.job_type else None,
            "status": str(work_order.status) if work_order.status else None,
            "priority": str(work_order.priority) if work_order.priority else None,
            "scheduled_date": work_order.scheduled_date.isoformat() if work_order.scheduled_date else None,
            "assigned_technician": work_order.assigned_technician,
        },
    )

    # Invalidate work order and dashboard caches
    await get_cache_service().delete_pattern("workorders:*")
    await get_cache_service().delete_pattern("dashboard:*")

    # Fetch customer for name population in response
    customer = None
    if work_order.customer_id:
        cust_result = await db.execute(
            select(Customer).where(Customer.id == work_order.customer_id)
        )
        customer = cust_result.scalar_one_or_none()

    return work_order_with_customer_name(work_order, customer)


@router.patch("/{work_order_id}", response_model=WorkOrderResponse)
async def update_work_order(
    work_order_id: str,
    work_order_data: WorkOrderUpdate,
    request: Request,
    db: DbSession,
    current_user: CurrentUser,
):
    """Update a work order."""
    # First check if work order exists
    result = await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))
    work_order = result.scalar_one_or_none()

    if not work_order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Work order not found",
        )

    update_data = work_order_data.model_dump(exclude_unset=True)

    if not update_data:
        # Still need customer name for response
        customer = None
        billing_cust = None
        if work_order.customer_id:
            cust_result = await db.execute(
                select(Customer).where(Customer.id == work_order.customer_id)
            )
            customer = cust_result.scalar_one_or_none()
        if work_order.billing_customer_id:
            bc_result = await db.execute(
                select(Customer).where(Customer.id == work_order.billing_customer_id)
            )
            billing_cust = bc_result.scalar_one_or_none()
        return work_order_with_customer_name(work_order, customer, billing_cust)

    # Track status change for WebSocket event
    old_status = str(work_order.status) if work_order.status else None
    old_technician = work_order.assigned_technician

    # Capture before-state for audit diff
    audit_changes = {}
    for field, new_value in update_data.items():
        old_value = getattr(work_order, field, None)
        # Serialize for JSON comparison
        old_ser = str(old_value) if old_value is not None else None
        new_ser = str(new_value) if new_value is not None else None
        if old_ser != new_ser:
            audit_changes[field] = {"old": old_ser, "new": new_ser}

    try:
        # Use SQLAlchemy ORM update - handles ENUM types correctly
        for field, value in update_data.items():
            setattr(work_order, field, value)

        # Auto-resolve assigned_technician (name string) → technician_id (UUID FK)
        # The schedule UI only sets assigned_technician, but the technician dashboard
        # needs technician_id to find jobs. Bridge the gap automatically.
        if "assigned_technician" in update_data and update_data["assigned_technician"]:
            tech_name = update_data["assigned_technician"]
            name_parts = tech_name.strip().split(None, 1)
            if len(name_parts) >= 2:
                tech_result = await db.execute(
                    select(Technician).where(
                        Technician.first_name == name_parts[0],
                        Technician.last_name == name_parts[1],
                        Technician.is_active == True,
                    ).limit(1)
                )
            else:
                tech_result = await db.execute(
                    select(Technician).where(
                        Technician.first_name == tech_name,
                        Technician.is_active == True,
                    ).limit(1)
                )
            matched_tech = tech_result.scalar_one_or_none()
            if matched_tech:
                work_order.technician_id = matched_tech.id
                logger.info(f"Auto-resolved technician '{tech_name}' → {matched_tech.id}")

        # Update timestamp and audit
        work_order.updated_at = datetime.utcnow()
        work_order.updated_by = current_user.email if current_user else None

        await db.commit()
        await db.refresh(work_order)

        # Create audit log entry for the update
        if audit_changes:
            try:
                # Determine action type from changes
                action = "updated"
                desc_parts = []
                if "status" in audit_changes:
                    action = "status_changed"
                    desc_parts.append(f"Status: {audit_changes['status']['old']} → {audit_changes['status']['new']}")
                if "assigned_technician" in audit_changes or "technician_id" in audit_changes:
                    action = "assigned" if "status" not in audit_changes else action
                    desc_parts.append(f"Technician: {audit_changes.get('assigned_technician', {}).get('old', '?')} → {audit_changes.get('assigned_technician', {}).get('new', '?')}")
                if not desc_parts:
                    desc_parts.append(f"Updated {', '.join(audit_changes.keys())}")

                client_ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else None)
                audit_entry = WorkOrderAuditLog(
                    work_order_id=work_order.id,
                    action=action,
                    description="; ".join(desc_parts),
                    user_email=current_user.email if current_user else None,
                    user_name=getattr(current_user, "full_name", None) or (current_user.email if current_user else "System"),
                    source=request.headers.get("X-Source", "crm"),
                    ip_address=client_ip,
                    user_agent=request.headers.get("User-Agent", None),
                    changes=audit_changes,
                )
                db.add(audit_entry)
                await db.commit()
            except Exception as audit_err:
                logger.warning(f"Failed to create audit log for WO {work_order_id}: {audit_err}")

    except Exception as e:
        await db.rollback()
        logger.error(f"Error updating work order {work_order_id}: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")

    # Outlook calendar sync — fire-and-forget when technician is assigned
    if ("assigned_technician" in update_data or "technician_id" in update_data) and work_order.technician_id:
        try:
            import asyncio
            from app.services.ms365_calendar_service import MS365CalendarService
            tech_result2 = await db.execute(
                select(Technician).where(Technician.id == work_order.technician_id)
            )
            tech = tech_result2.scalar_one_or_none()
            ms_email = getattr(tech, "microsoft_email", None) if tech else None
            if ms_email and MS365CalendarService.is_configured():
                customer2 = None
                if work_order.customer_id:
                    cr = await db.execute(select(Customer).where(Customer.id == work_order.customer_id))
                    customer2 = cr.scalar_one_or_none()
                cust_name = f"{customer2.first_name} {customer2.last_name}" if customer2 else "Unknown"
                job = str(work_order.job_type) if work_order.job_type else "Service"
                addr = work_order.service_address_line1 or ""
                sched = work_order.scheduled_date.isoformat() if work_order.scheduled_date else ""
                t_start = work_order.time_window_start
                t_end = work_order.time_window_end

                from datetime import datetime as dt2, time as time2
                # Build start datetime from scheduled_date + time_window_start
                if sched and t_start:
                    start_datetime = dt2.combine(work_order.scheduled_date, t_start)
                elif sched:
                    start_datetime = dt2.fromisoformat(f"{sched}T08:00:00")
                else:
                    start_datetime = dt2.now()
                dur = work_order.estimated_duration_hours or 2.0
                wo_id = str(work_order.id)
                wo_notes = work_order.notes or ""
                existing_event_id = work_order.outlook_event_id

                tech_name_for_shared = tech.first_name + " " + tech.last_name if tech else "Unassigned"
                existing_shared_event_id = getattr(work_order, "outlook_shared_event_id", None)

                async def _sync_calendar():
                    try:
                        if existing_event_id:
                            await MS365CalendarService.update_event(
                                technician_microsoft_email=ms_email,
                                event_id=existing_event_id,
                                subject=f"{job.title()} - {cust_name}",
                                location=addr,
                                body=f"Work Order: {wo_id}\nNotes: {wo_notes}",
                                start_dt=start_datetime,
                                duration_hours=dur,
                            )
                        else:
                            new_event_id = await MS365CalendarService.create_event(
                                technician_microsoft_email=ms_email,
                                subject=f"{job.title()} - {cust_name}",
                                location=addr,
                                body=f"Work Order: {wo_id}\nNotes: {wo_notes}",
                                start_dt=start_datetime,
                                duration_hours=dur,
                            )
                            if new_event_id:
                                from app.database import async_session_maker
                                async with async_session_maker() as sess:
                                    from sqlalchemy import text as sql_text
                                    await sess.execute(
                                        sql_text("UPDATE work_orders SET outlook_event_id = :eid WHERE id = :wid"),
                                        {"eid": new_event_id, "wid": wo_id},
                                    )
                                    await sess.commit()
                                logger.info(f"Outlook event {new_event_id} created for WO {wo_id}")
                    except Exception as cal_err:
                        logger.warning(f"Outlook calendar sync failed for WO {wo_id}: {cal_err}")

                    # Shared mailbox calendar — independent try/except, never blocks individual sync
                    try:
                        if MS365CalendarService.shared_calendar_configured():
                            shared_subject = MS365CalendarService.build_shared_event_subject(job, cust_name, tech_name_for_shared)
                            shared_body = f"Work Order: {wo_id}\nTech: {tech_name_for_shared}\nNotes: {wo_notes}"
                            if existing_shared_event_id:
                                await MS365CalendarService.update_shared_event(
                                    event_id=existing_shared_event_id,
                                    subject=shared_subject,
                                    location=addr,
                                    body=shared_body,
                                    start_dt=start_datetime,
                                    duration_hours=dur,
                                )
                            else:
                                new_shared_id = await MS365CalendarService.create_shared_event(
                                    subject=shared_subject,
                                    location=addr,
                                    body=shared_body,
                                    start_dt=start_datetime,
                                    duration_hours=dur,
                                )
                                if new_shared_id:
                                    from app.database import async_session_maker as asm2
                                    async with asm2() as sess2:
                                        from sqlalchemy import text as sql_text2
                                        await sess2.execute(
                                            sql_text2("UPDATE work_orders SET outlook_shared_event_id = :eid WHERE id = :wid"),
                                            {"eid": new_shared_id, "wid": wo_id},
                                        )
                                        await sess2.commit()
                                    logger.info(f"Shared calendar event {new_shared_id} created for WO {wo_id}")
                    except Exception as shared_err:
                        logger.warning(f"Shared calendar sync failed for WO {wo_id}: {shared_err}")

                asyncio.create_task(_sync_calendar())
        except Exception as e:
            logger.warning(f"Calendar sync setup error: {e}")

    # When status changes to "canceled", delete both calendar events
    if old_status != str(work_order.status) and str(work_order.status) == "canceled":
        try:
            import asyncio as asyncio2
            from app.services.ms365_calendar_service import MS365CalendarService as CalSvc

            async def _cancel_calendar_events():
                try:
                    # Delete individual technician event
                    if work_order.outlook_event_id and work_order.technician_id:
                        tr = await db.execute(select(Technician).where(Technician.id == work_order.technician_id))
                        t = tr.scalar_one_or_none()
                        t_email = getattr(t, "microsoft_email", None) if t else None
                        if t_email:
                            await CalSvc.delete_event(t_email, work_order.outlook_event_id)
                    # Delete shared calendar event
                    shared_eid = getattr(work_order, "outlook_shared_event_id", None)
                    if shared_eid and CalSvc.shared_calendar_configured():
                        await CalSvc.delete_shared_event(shared_eid)
                    # Clear event IDs
                    from app.database import async_session_maker as asm_cancel
                    async with asm_cancel() as sess_cancel:
                        from sqlalchemy import text as sql_cancel
                        await sess_cancel.execute(
                            sql_cancel("UPDATE work_orders SET outlook_event_id = NULL, outlook_shared_event_id = NULL WHERE id = :wid"),
                            {"wid": str(work_order.id)},
                        )
                        await sess_cancel.commit()
                except Exception as cancel_err:
                    logger.warning(f"Calendar cancel cleanup failed for WO {work_order.id}: {cancel_err}")

            asyncio2.create_task(_cancel_calendar_events())
        except Exception as e:
            logger.warning(f"Calendar cancel setup error: {e}")

    # Invalidate caches
    await get_cache_service().delete_pattern("workorders:*")
    await get_cache_service().delete_pattern("dashboard:*")

    # Fetch customer + billing customer for name population in response
    customer = None
    billing_cust = None
    if work_order.customer_id:
        cust_result = await db.execute(
            select(Customer).where(Customer.id == work_order.customer_id)
        )
        customer = cust_result.scalar_one_or_none()
    if work_order.billing_customer_id:
        bc_result = await db.execute(
            select(Customer).where(Customer.id == work_order.billing_customer_id)
        )
        billing_cust = bc_result.scalar_one_or_none()

    # Determine the type of update for WebSocket event
    new_status = str(work_order.status) if work_order.status else None
    new_technician = work_order.assigned_technician

    # Broadcast appropriate WebSocket events
    event_data = {
        "id": work_order.id,
        "customer_id": str(work_order.customer_id),
        "job_type": str(work_order.job_type) if work_order.job_type else None,
        "status": new_status,
        "priority": str(work_order.priority) if work_order.priority else None,
        "scheduled_date": work_order.scheduled_date.isoformat() if work_order.scheduled_date else None,
        "assigned_technician": new_technician,
        "updated_fields": list(update_data.keys()),
    }

    # Auto-generate invoice when status changes to "completed" via PATCH
    if old_status != new_status and new_status == "completed":
        try:
            from app.models.invoice import Invoice
            from datetime import timedelta as td

            existing_inv = await db.execute(
                select(Invoice).where(Invoice.work_order_id == work_order_id)
            )
            if not existing_inv.scalar_one_or_none():
                wo_amount = float(work_order.total_amount) if work_order.total_amount else 0.0
                if wo_amount > 0 and work_order.customer_id:
                    job_labels = {
                        "pumping": "Septic Tank Pumping", "inspection": "Septic System Inspection",
                        "real_estate_inspection": "Real Estate Inspection",
                        "repair": "Septic System Repair", "installation": "Septic System Installation",
                        "emergency": "Emergency Service Call", "maintenance": "Septic Maintenance",
                        "grease_trap": "Grease Trap Service", "camera_inspection": "Camera Inspection",
                    }
                    job_label = job_labels.get(str(work_order.job_type) if work_order.job_type else "pumping", "Septic Service")
                    addr_parts = [work_order.service_address_line1, work_order.service_city, work_order.service_state]
                    addr = ", ".join(p for p in addr_parts if p)
                    desc = job_label + (f" at {addr}" if addr else "")
                    tax_rate, tax = 8.25, round(wo_amount * 0.0825, 2)
                    total = round(wo_amount + tax, 2)
                    inv_customer_id = work_order.billing_customer_id or work_order.customer_id
                    inv = Invoice(
                        id=uuid.uuid4(), customer_id=inv_customer_id, work_order_id=work_order.id,
                        invoice_number=f"INV-{datetime.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:4].upper()}",
                        issue_date=datetime.now().date(), due_date=datetime.now().date() + td(days=30),
                        amount=total, paid_amount=0, status="draft",
                        line_items=[{"description": desc, "quantity": 1, "unit_price": wo_amount, "amount": wo_amount}],
                        notes=f"Auto-generated from {work_order.work_order_number or 'work order'} completion",
                    )
                    db.add(inv)
                    await db.commit()
                    logger.info(f"Auto-generated invoice {inv.invoice_number} for WO {work_order_id} (via PATCH)")
                    # Send "Pay Now" SMS with customer portal link — PAUSED 2026-04-03
                    # Re-enable when ready for real-time invoice notifications
                    # try:
                    #     if customer and customer.phone:
                    #         from app.services.sms_service import sms_service as sms_svc
                    #         pay_url = f"https://react.ecbtx.com/portal/pay/{inv.id}"
                    #         sms_msg = (
                    #             f"Hi {customer.first_name or 'there'}, your invoice "
                    #             f"#{inv.invoice_number} for ${float(total):.2f} is ready. "
                    #             f"Pay online: {pay_url}"
                    #         )
                    #         await sms_svc.send_sms(customer.phone, sms_msg)
                    #         logger.info(f"Sent Pay Now SMS to {customer.phone} for invoice {inv.invoice_number}")
                    # except Exception as sms_err:
                    #     logger.warning(f"Pay Now SMS failed for invoice {inv.invoice_number}: {sms_err}")
                    logger.info(f"Invoice SMS paused — skipping Pay Now SMS for {inv.invoice_number}")
        except Exception as e:
            await db.rollback()
            logger.warning(f"Auto-invoice generation failed for WO {work_order_id}: {e}")

    # Auto-notify customer via SMS when job is marked completed — PAUSED 2026-04-03
    # Re-enable when ready for real-time completion notifications
    notification_sent = False
    # if old_status != new_status and new_status == "completed":
    #     try:
    #         from app.services.sms_service import sms_service as sms_svc
    #         if work_order.customer_id and customer:
    #             phone = customer.phone
    #             if phone:
    #                 addr_parts = [work_order.service_address_line1, work_order.service_city]
    #                 addr = ", ".join(p for p in addr_parts if p) or "your property"
    #                 msg = (
    #                     f"Hi {customer.first_name}! Your septic service at {addr} is complete. "
    #                     f"Thank you for choosing MAC Septic. Questions? Call (512) 353-0555."
    #                 )
    #                 if sms_svc.is_configured:
    #                     await sms_svc.send_sms(to=phone, body=msg)
    #                     notification_sent = True
    #                     logger.info(f"Completion SMS sent to {phone} for WO {work_order_id}")
    #                 else:
    #                     logger.info("SMS service not configured — completion SMS skipped")
    #     except Exception as e:
    #         logger.warning(f"Completion SMS failed for WO {work_order_id}: {e}")
    if old_status != new_status and new_status == "completed":
        logger.info(f"Completion SMS paused — skipping notification for WO {work_order_id}")

    # Auto-email real estate inspection report to Doug on completion
    if old_status != new_status and new_status == "completed":
        await _send_real_estate_inspection_report(work_order, db)

    # Auto-upload offline conversion to Google Ads (fire-and-forget)
    if old_status != new_status and new_status == "completed":
        try:
            from app.services.google_ads_service import get_google_ads_service
            ads_svc = get_google_ads_service()
            if ads_svc.is_configured() and ads_svc.conversion_action_id:
                wo_amount = float(work_order.total_amount) if work_order.total_amount else 700.0
                cust_phone = customer.phone if customer else None
                cust_email = customer.email if customer else None
                if cust_phone or cust_email:
                    upload_result = await ads_svc.upload_offline_conversion(
                        phone=cust_phone,
                        email=cust_email,
                        conversion_value=wo_amount,
                        conversion_time=datetime.utcnow(),
                        order_id=str(work_order_id),
                    )
                    if upload_result.get("success"):
                        logger.info(f"Google Ads offline conversion uploaded for WO {work_order_id}: ${wo_amount}")
                    else:
                        logger.warning(f"Google Ads conversion upload failed for WO {work_order_id}: {upload_result.get('error')}")
        except Exception as e:
            logger.warning(f"Google Ads conversion upload error for WO {work_order_id}: {e}")

    # Status change event
    if old_status != new_status:
        await manager.broadcast_event(
            event_type="job_status",
            data={
                **event_data,
                "old_status": old_status,
                "new_status": new_status,
            },
        )

    # Technician assignment event
    if old_technician != new_technician:
        await manager.broadcast_event(
            event_type="dispatch_update",
            data={
                **event_data,
                "old_technician": old_technician,
                "new_technician": new_technician,
            },
        )

    # General update event (always sent)
    await manager.broadcast_event(
        event_type="work_order_update",
        data=event_data,
    )

    # Schedule change event (when schedule-related fields change)
    schedule_fields = {"scheduled_date", "time_window_start", "time_window_end", "assigned_technician"}
    if schedule_fields.intersection(update_data.keys()):
        await manager.broadcast_event(
            event_type="schedule_change",
            data={
                "work_order_id": work_order.id,
                "customer_id": str(work_order.customer_id),
                "scheduled_date": work_order.scheduled_date.isoformat() if work_order.scheduled_date else None,
                "time_window_start": str(work_order.time_window_start) if work_order.time_window_start else None,
                "time_window_end": str(work_order.time_window_end) if work_order.time_window_end else None,
                "assigned_technician": new_technician,
                "updated_fields": list(schedule_fields.intersection(update_data.keys())),
            },
        )

    response_data = work_order_with_customer_name(work_order, customer, billing_cust)
    response_data["notification_sent"] = notification_sent
    return response_data


@router.delete("/{work_order_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_work_order(
    work_order_id: str,
    db: DbSession,
    current_user: CurrentUser,
):
    """Delete a work order."""
    try:
        result = await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))
        work_order = result.scalar_one_or_none()

        if not work_order:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Work order not found",
            )

        # Fire-and-forget cleanup of calendar events before deletion
        try:
            from app.services.ms365_calendar_service import MS365CalendarService
            if MS365CalendarService.is_configured():
                if work_order.outlook_event_id and work_order.technician_id:
                    tech_result = await db.execute(
                        select(Technician).where(Technician.id == work_order.technician_id)
                    )
                    tech = tech_result.scalar_one_or_none()
                    ms_email = getattr(tech, "microsoft_email", None) if tech else None
                    if ms_email:
                        try:
                            await MS365CalendarService.delete_event(ms_email, work_order.outlook_event_id)
                        except Exception:
                            pass
                shared_eid = getattr(work_order, "outlook_shared_event_id", None)
                if shared_eid and MS365CalendarService.shared_calendar_configured():
                    try:
                        await MS365CalendarService.delete_shared_event(shared_eid)
                    except Exception:
                        pass
        except Exception as cal_err:
            logging.warning(f"Calendar cleanup failed for WO {work_order_id}: {cal_err}")

        # Nullify FK references in related tables before deletion
        # Each cleanup runs in its own transaction to avoid cascade abort
        from sqlalchemy import text as sql_text
        from app.database import async_session_maker
        wo_id = str(work_order.id)

        async def _cleanup_sql(query: str):
            async with async_session_maker() as s:
                try:
                    await s.execute(sql_text(query), {"wid": wo_id})
                    await s.commit()
                except Exception:
                    await s.rollback()

        # SET NULL on nullable FKs
        for tbl, col in [
            ("bookings", "work_order_id"),
            ("invoices", "work_order_id"),
            ("payments", "work_order_id"),
            ("quotes", "converted_to_work_order_id"),
            ("tickets", "work_order_id"),
            ("gps_locations", "current_work_order_id"),
            ("gps_breadcrumbs", "work_order_id"),
            ("gps_geofence_events", "work_order_id"),
            ("gps_eta_estimates", "work_order_id"),
        ]:
            await _cleanup_sql(f"UPDATE {tbl} SET {col} = NULL WHERE {col} = :wid")

        # DELETE cascade-eligible child rows
        for tbl in [
            "work_order_audit_log", "work_order_photos", "job_costs",
            "gps_work_order_tracking", "gps_route_history",
        ]:
            await _cleanup_sql(f"DELETE FROM {tbl} WHERE work_order_id = :wid")

        await db.delete(work_order)
        await db.commit()
        await get_cache_service().delete_pattern("workorders:*")
        await get_cache_service().delete_pattern("dashboard:*")
    except HTTPException:
        raise
    except Exception as e:
        import logging
        import traceback
        logging.error(f"Error deleting work order {work_order_id}: {e}")
        logging.error(traceback.format_exc())
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete work order: {str(e)}",
        )


class WorkOrderCompleteRequest(BaseModel):
    """Request to complete a work order."""

    dump_site_id: Optional[str] = None
    notes: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None


@router.post("/{work_order_id}/complete")
async def complete_work_order(
    work_order_id: str,
    request: WorkOrderCompleteRequest,
    db: DbSession,
    current_user: CurrentUser,
):
    """Complete a work order and auto-create commission.

    This endpoint marks a work order as completed and automatically creates
    a commission record for the assigned technician based on job type and
    configured commission rates.

    For pumping and grease_trap jobs, a dump_site_id should be provided to
    calculate dump fee deductions from the commission.
    """
    result = await db.execute(select(WorkOrder).where(WorkOrder.id == work_order_id))
    work_order = result.scalar_one_or_none()

    if not work_order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Work order not found",
        )

    if str(work_order.status) == "completed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Work order is already completed",
        )

    # Use ORM setattr like the PATCH endpoint does - handles ENUM types correctly
    work_order.status = "completed"
    work_order.actual_end_time = datetime.now()
    work_order.updated_at = datetime.utcnow()

    if request.latitude and request.longitude:
        work_order.clock_out_gps_lat = request.latitude
        work_order.clock_out_gps_lon = request.longitude

    if request.notes:
        existing_notes = work_order.notes or ""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        work_order.notes = f"{existing_notes}\n[{timestamp}] Completion: {request.notes}".strip()

    # Calculate labor minutes if start time exists
    if work_order.actual_start_time:
        duration = datetime.now() - work_order.actual_start_time
        work_order.total_labor_minutes = int(duration.total_seconds() / 60)

    # Commit the status change FIRST, before commission
    try:
        await db.commit()
        await db.refresh(work_order)
        logger.info(f"Committed work order {work_order_id} status to completed")
    except Exception as e:
        await db.rollback()
        logger.error(f"Failed to commit work order completion: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to complete work order: {str(e)}")

    # Now auto-create commission (in separate transaction)
    commission = await auto_create_commission(
        db=db,
        work_order=work_order,
        dump_site_id=request.dump_site_id,
    )

    # Commit commission if created
    if commission:
        try:
            await db.commit()
            logger.info(f"Created commission {commission.id} for work order {work_order_id}")
        except Exception as e:
            await db.rollback()
            logger.error(f"Failed to create commission: {e}")
            commission = None  # Still return success for work order completion

    # Auto-generate invoice (non-blocking, won't fail WO completion)
    invoice_info = None
    try:
        from app.models.invoice import Invoice
        from datetime import timedelta as td

        # Check if invoice already exists for this WO
        existing_inv = await db.execute(
            select(Invoice).where(Invoice.work_order_id == work_order_id)
        )
        if not existing_inv.scalar_one_or_none():
            wo_amount = float(work_order.total_amount) if work_order.total_amount else 0.0
            if wo_amount > 0 and work_order.customer_id:
                job_labels = {
                    "pumping": "Septic Tank Pumping",
                    "inspection": "Septic System Inspection",
                    "repair": "Septic System Repair",
                    "installation": "Septic System Installation",
                    "emergency": "Emergency Service Call",
                    "maintenance": "Septic Maintenance",
                    "grease_trap": "Grease Trap Service",
                    "camera_inspection": "Camera Inspection",
                }
                job_label = job_labels.get(
                    str(work_order.job_type) if work_order.job_type else "pumping",
                    "Septic Service",
                )
                addr_parts = [work_order.service_address_line1, work_order.service_city, work_order.service_state]
                addr = ", ".join(p for p in addr_parts if p)
                desc = job_label
                if addr:
                    desc += f" at {addr}"

                tax_rate = 8.25
                tax = round(wo_amount * tax_rate / 100, 2)
                total = round(wo_amount + tax, 2)
                date_part = datetime.now().strftime("%Y%m%d")
                random_part = uuid.uuid4().hex[:4].upper()

                invoice = Invoice(
                    id=uuid.uuid4(),
                    customer_id=work_order.customer_id,
                    work_order_id=work_order.id,
                    invoice_number=f"INV-{date_part}-{random_part}",
                    issue_date=datetime.now().date(),
                    due_date=datetime.now().date() + td(days=30),
                    amount=total,
                    paid_amount=0,
                    status="draft",
                    line_items=[{"description": desc, "quantity": 1, "unit_price": wo_amount, "amount": wo_amount}],
                    notes=f"Auto-generated from {work_order.work_order_number or 'work order'} completion",
                )
                db.add(invoice)
                await db.commit()
                await db.refresh(invoice)
                invoice_info = {"id": str(invoice.id), "invoice_number": invoice.invoice_number, "total": total}
                logger.info(f"Auto-generated invoice {invoice.invoice_number} for WO {work_order_id}")
    except Exception as e:
        await db.rollback()
        logger.warning(f"Auto-invoice generation failed for WO {work_order_id}: {e}")

    # Store values before any potential errors
    wo_id = work_order.id
    wo_number = work_order.work_order_number
    labor_mins = work_order.total_labor_minutes
    tech_id = work_order.technician_id
    cust_id = work_order.customer_id
    comm_id = str(commission.id) if commission else None
    comm_amount = float(commission.commission_amount) if commission else None
    comm_status = commission.status if commission else None
    comm_job_type = commission.job_type if commission else None
    comm_rate = commission.rate if commission else None

    # Get customer name (in try block to not affect commit)
    customer_name = None
    try:
        if cust_id:
            cust_result = await db.execute(select(Customer).where(Customer.id == cust_id))
            customer = cust_result.scalar_one_or_none()
            if customer:
                customer_name = f"{customer.first_name or ''} {customer.last_name or ''}".strip()
    except Exception as e:
        logger.warning(f"Failed to get customer name: {e}")

    # Broadcast WebSocket event (non-blocking, errors don't affect response)
    try:
        await manager.broadcast_event(
            event_type="job_status",
            data={
                "id": wo_id,
                "status": "completed",
                "technician_id": tech_id,
                "commission_id": comm_id,
                "commission_amount": comm_amount,
            },
        )
    except Exception as e:
        logger.warning(f"Failed to broadcast WebSocket event: {e}")

    # Auto-email real estate inspection report to Doug on completion
    await _send_real_estate_inspection_report(work_order, db)

    return {
        "id": wo_id,
        "work_order_number": wo_number,
        "status": "completed",
        "customer_name": customer_name,
        "labor_minutes": labor_mins,
        "commission": {
            "id": comm_id,
            "amount": comm_amount,
            "status": comm_status,
            "job_type": comm_job_type,
            "rate": comm_rate,
        }
        if commission
        else None,
        "invoice": invoice_info,
    }


# =====================================================
# Invoice Generation from Work Order
# =====================================================


@router.post("/{work_order_id}/generate-invoice")
async def generate_invoice_from_work_order(
    work_order_id: str,
    db: DbSession,
    current_user: CurrentUser,
):
    """Auto-generate an invoice from a completed work order.

    Creates an invoice with line items derived from the work order's
    job type, total_amount, and service details. Sets net-30 payment terms.
    """
    from app.models.invoice import Invoice
    from datetime import timedelta

    # Fetch work order with customer
    result = await db.execute(
        select(WorkOrder, Customer)
        .outerjoin(Customer, WorkOrder.customer_id == Customer.id)
        .where(WorkOrder.id == work_order_id)
    )
    row = result.first()

    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Work order not found",
        )

    work_order, customer = row

    # Check if invoice already exists for this work order
    existing = await db.execute(
        select(Invoice).where(Invoice.work_order_id == work_order_id)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="An invoice already exists for this work order",
        )

    # Build line items from work order data
    job_type_labels = {
        "pumping": "Septic Tank Pumping",
        "inspection": "Septic System Inspection",
        "repair": "Septic System Repair",
        "installation": "Septic System Installation",
        "emergency": "Emergency Service Call",
        "maintenance": "Septic Maintenance",
        "grease_trap": "Grease Trap Service",
        "camera_inspection": "Camera Inspection",
    }

    job_label = job_type_labels.get(
        str(work_order.job_type) if work_order.job_type else "pumping",
        "Septic Service",
    )
    amount = float(work_order.total_amount) if work_order.total_amount else 0.0

    # Build description from service address
    address_parts = [
        work_order.service_address_line1,
        work_order.service_city,
        work_order.service_state,
    ]
    address = ", ".join(p for p in address_parts if p)
    description = f"{job_label}"
    if address:
        description += f" at {address}"
    if work_order.scheduled_date:
        description += f" on {work_order.scheduled_date}"
    if work_order.estimated_gallons:
        description += f" ({work_order.estimated_gallons} gallons)"

    line_items = [
        {
            "description": description,
            "quantity": 1,
            "unit_price": amount,
            "amount": amount,
        }
    ]

    # Calculate tax (8.25% default)
    tax_rate = 8.25
    subtotal = amount
    tax = round(subtotal * tax_rate / 100, 2)
    total = round(subtotal + tax, 2)

    # Generate invoice number
    date_part = datetime.now().strftime("%Y%m%d")
    random_part = uuid.uuid4().hex[:4].upper()
    invoice_number = f"INV-{date_part}-{random_part}"

    # Create invoice — bill to billing_customer if set, otherwise service customer
    today = datetime.now().date()
    invoice_customer_id = work_order.billing_customer_id or work_order.customer_id
    invoice = Invoice(
        id=uuid.uuid4(),
        customer_id=invoice_customer_id,
        work_order_id=work_order.id,
        invoice_number=invoice_number,
        issue_date=today,
        due_date=today + timedelta(days=30),
        amount=total,
        paid_amount=0,
        status="draft",
        line_items=line_items,
        notes=f"Generated from {work_order.work_order_number or 'work order'}",
    )

    db.add(invoice)
    await db.commit()
    await db.refresh(invoice)

    # Invalidate caches
    await get_cache_service().delete_pattern("dashboard:*")

    customer_name = None
    if customer:
        customer_name = f"{customer.first_name or ''} {customer.last_name or ''}".strip()

    return {
        "id": str(invoice.id),
        "invoice_number": invoice.invoice_number,
        "customer_id": str(invoice.customer_id),
        "customer_name": customer_name,
        "work_order_id": str(invoice.work_order_id),
        "work_order_number": work_order.work_order_number,
        "issue_date": invoice.issue_date.isoformat(),
        "due_date": invoice.due_date.isoformat(),
        "subtotal": subtotal,
        "tax_rate": tax_rate,
        "tax": tax,
        "total": total,
        "status": invoice.status,
        "line_items": line_items,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Workorder Estimate Endpoints
#
# A workorder estimate is a Quote row with kind='wo_estimate' attached to a
# specific work_order_id. It can be edited freely while in 'draft' or 'sent'
# state, then converted into an Invoice (which is pushed to QuickBooks via
# qbo_service.sync_invoice). Each work order has at most one ACTIVE
# (status != 'converted') estimate at a time.
# ──────────────────────────────────────────────────────────────────────────────

from app.models.quote import Quote as _EstimateQuote
from app.schemas.types import UUIDStr as _EstimateUUIDStr


class _EstimateLineItem(BaseModel):
    description: str = ""
    qty: float = 0.0
    unit_price: float = 0.0
    line_total: Optional[float] = None


class EstimateCreate(BaseModel):
    line_items: Optional[List[_EstimateLineItem]] = None
    notes: Optional[str] = None
    tax_rate: Optional[float] = 0.0


class EstimateUpdate(BaseModel):
    line_items: Optional[List[_EstimateLineItem]] = None
    tax_rate: Optional[float] = None
    notes: Optional[str] = None
    status: Optional[str] = None  # only 'draft' or 'sent' allowed here


class EstimateOut(BaseModel):
    id: _EstimateUUIDStr
    work_order_id: Optional[_EstimateUUIDStr] = None
    customer_id: Optional[_EstimateUUIDStr] = None
    kind: str
    status: str
    line_items: List[_EstimateLineItem] = Field(default_factory=list)
    subtotal: float = 0.0
    tax_rate: float = 0.0
    tax: float = 0.0
    total: float = 0.0
    notes: Optional[str] = None
    converted_to_invoice_id: Optional[_EstimateUUIDStr] = None
    converted_at: Optional[datetime] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None


class ConvertResult(BaseModel):
    invoice_id: _EstimateUUIDStr
    qbo_invoice_id: Optional[str] = None
    qbo_pushed: bool


def _normalize_line_items(raw_items) -> List[dict]:
    """Coerce incoming line items to the canonical dict shape and
    recompute line_total defensively as qty * unit_price."""
    out: List[dict] = []
    if not raw_items:
        return out
    for item in raw_items:
        if isinstance(item, _EstimateLineItem):
            data = item.model_dump()
        elif isinstance(item, dict):
            data = dict(item)
        else:
            continue
        try:
            qty = float(data.get("qty") or 0)
        except (TypeError, ValueError):
            qty = 0.0
        try:
            unit_price = float(data.get("unit_price") or 0)
        except (TypeError, ValueError):
            unit_price = 0.0
        out.append(
            {
                "description": str(data.get("description") or ""),
                "qty": qty,
                "unit_price": unit_price,
                "line_total": round(qty * unit_price, 2),
            }
        )
    return out


def _compute_estimate_totals(items: List[dict], tax_rate: float):
    subtotal = sum(float(i.get("line_total") or 0) for i in items)
    rate = float(tax_rate or 0)
    tax = round(subtotal * rate / 100, 2) if rate else 0.0
    total = round(subtotal + tax, 2)
    return round(subtotal, 2), tax, total


def _serialize_estimate(q: _EstimateQuote) -> EstimateOut:
    return EstimateOut(
        id=q.id,
        work_order_id=q.work_order_id,
        customer_id=q.customer_id,
        kind=q.kind or "wo_estimate",
        status=q.status or "draft",
        line_items=_normalize_line_items(q.line_items or []),
        subtotal=float(q.subtotal or 0),
        tax_rate=float(q.tax_rate or 0),
        tax=float(q.tax or 0),
        total=float(q.total or 0),
        notes=q.notes,
        converted_to_invoice_id=q.converted_to_invoice_id,
        converted_at=q.converted_at,
        created_at=q.created_at,
        updated_at=q.updated_at,
    )


async def _get_active_estimate(db, wo_id: str) -> Optional[_EstimateQuote]:
    res = await db.execute(
        select(_EstimateQuote).where(
            _EstimateQuote.work_order_id == wo_id,
            _EstimateQuote.kind == "wo_estimate",
            _EstimateQuote.status != "converted",
        )
    )
    return res.scalars().first()


@router.get("/{wo_id}/estimate")
async def get_workorder_estimate(
    wo_id: str,
    db: DbSession,
    current_user: CurrentUser,
):
    """Return the active wo_estimate quote for this work order, or null."""
    est = await _get_active_estimate(db, wo_id)
    if not est:
        return None
    return _serialize_estimate(est)


@router.post(
    "/{wo_id}/estimate",
    response_model=EstimateOut,
    status_code=status.HTTP_201_CREATED,
)
async def create_workorder_estimate(
    wo_id: str,
    payload: EstimateCreate,
    db: DbSession,
    current_user: CurrentUser,
):
    """Create a new wo_estimate for this work order. 409 if one already active."""
    # Fetch work order
    res = await db.execute(select(WorkOrder).where(WorkOrder.id == wo_id))
    wo = res.scalars().first()
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")

    # Reject if active estimate already exists
    existing = await _get_active_estimate(db, wo_id)
    if existing:
        raise HTTPException(
            status_code=409,
            detail="An active estimate already exists for this work order",
        )

    items = _normalize_line_items(payload.line_items)
    tax_rate = float(payload.tax_rate or 0)
    subtotal, tax, total = _compute_estimate_totals(items, tax_rate)

    # Generate quote number
    date_part = datetime.now().strftime("%Y%m%d")
    rand_part = uuid.uuid4().hex[:4].upper()
    quote_number = f"EST-{date_part}-{rand_part}"

    quote = _EstimateQuote(
        id=uuid.uuid4(),
        quote_number=quote_number,
        customer_id=wo.customer_id,
        work_order_id=wo.id,
        kind="wo_estimate",
        status="draft",
        line_items=items,
        subtotal=subtotal,
        tax_rate=tax_rate,
        tax=tax,
        total=total,
        notes=payload.notes,
    )
    db.add(quote)
    await db.commit()
    await db.refresh(quote)
    return _serialize_estimate(quote)


@router.patch("/{wo_id}/estimate", response_model=EstimateOut)
async def update_workorder_estimate(
    wo_id: str,
    payload: EstimateUpdate,
    db: DbSession,
    current_user: CurrentUser,
):
    """Update the active wo_estimate. Status may only be set to draft or sent here."""
    est = await _get_active_estimate(db, wo_id)
    if not est:
        raise HTTPException(status_code=404, detail="No active estimate for this work order")

    if payload.line_items is not None:
        est.line_items = _normalize_line_items(payload.line_items)
    if payload.tax_rate is not None:
        est.tax_rate = float(payload.tax_rate or 0)
    if payload.notes is not None:
        est.notes = payload.notes
    if payload.status is not None:
        if payload.status not in ("draft", "sent"):
            raise HTTPException(
                status_code=400,
                detail="Status may only be set to 'draft' or 'sent' here; use convert-to-invoice to mark converted",
            )
        est.status = payload.status

    # Always recompute totals server-side
    items = _normalize_line_items(est.line_items or [])
    est.line_items = items
    subtotal, tax, total = _compute_estimate_totals(items, float(est.tax_rate or 0))
    est.subtotal = subtotal
    est.tax = tax
    est.total = total

    await db.commit()
    await db.refresh(est)
    return _serialize_estimate(est)


@router.post("/{wo_id}/estimate/convert-to-invoice", response_model=ConvertResult)
async def convert_workorder_estimate_to_invoice(
    wo_id: str,
    db: DbSession,
    current_user: CurrentUser,
):
    """Convert the active wo_estimate to an Invoice and push to QuickBooks."""
    from app.models.invoice import Invoice
    from datetime import timedelta
    from app.services.qbo_service import get_qbo_service

    est = await _get_active_estimate(db, wo_id)
    if not est:
        raise HTTPException(status_code=404, detail="No active estimate for this work order")

    items = _normalize_line_items(est.line_items or [])
    valid_items = [i for i in items if i["qty"] > 0 and i["unit_price"] > 0]
    if not valid_items:
        raise HTTPException(
            status_code=400,
            detail="Estimate must have at least one line item with qty>0 and price>0",
        )

    # Recompute totals once more so the invoice numbers are authoritative
    subtotal, tax, total = _compute_estimate_totals(items, float(est.tax_rate or 0))

    # Generate invoice number
    date_part = datetime.now().strftime("%Y%m%d")
    rand_part = uuid.uuid4().hex[:4].upper()
    invoice_number = f"INV-{date_part}-{rand_part}"

    today = datetime.now().date()
    invoice = Invoice(
        id=uuid.uuid4(),
        customer_id=est.customer_id,
        work_order_id=est.work_order_id,
        invoice_number=invoice_number,
        issue_date=today,
        due_date=today + timedelta(days=30),
        amount=total,
        paid_amount=0,
        status="draft",
        line_items=items,
        notes=est.notes,
    )
    db.add(invoice)
    await db.flush()  # need invoice.id available before QBO push

    # Push to QBO — never fail the endpoint on QBO error
    qbo_pushed = False
    qbo_invoice_id: Optional[str] = None
    try:
        qbo = get_qbo_service()
        qbo_result = await qbo.sync_invoice(db, str(invoice.id))
        if qbo_result:
            qbo_pushed = True
            # QBO returns {"Invoice": {"Id": "...", ...}} or similar; be defensive
            if isinstance(qbo_result, dict):
                inv_obj = qbo_result.get("Invoice") or qbo_result
                if isinstance(inv_obj, dict):
                    qbo_invoice_id = (
                        inv_obj.get("Id")
                        or inv_obj.get("id")
                        or inv_obj.get("DocNumber")
                    )
            if qbo_invoice_id:
                invoice.quickbooks_invoice_id = str(qbo_invoice_id)
    except Exception:
        logger.exception("QBO sync failed for invoice %s", invoice.id)
        qbo_pushed = False

    # Mark estimate converted
    est.status = "converted"
    est.converted_to_invoice_id = invoice.id
    est.converted_at = datetime.utcnow()

    await db.commit()
    await db.refresh(invoice)

    return ConvertResult(
        invoice_id=invoice.id,
        qbo_invoice_id=qbo_invoice_id,
        qbo_pushed=qbo_pushed,
    )


# ──────────────────────────────────────────────────────────────────────────────
# Audit Log Endpoint
# ──────────────────────────────────────────────────────────────────────────────

@router.get("/{work_order_id}/audit-log", response_model=list[WorkOrderAuditLogResponse])
async def get_work_order_audit_log(
    work_order_id: str,
    db: DbSession,
    current_user: CurrentUser,
):
    """Get the full audit trail for a work order (newest first)."""
    result = await db.execute(
        select(WorkOrderAuditLog)
        .where(WorkOrderAuditLog.work_order_id == work_order_id)
        .order_by(desc(WorkOrderAuditLog.created_at))
    )
    entries = result.scalars().all()
    return entries


# ──────────────────────────────────────────────────────────────────────────────
# FOLLOW-UP SMS ("Book Today $25 Off")
# ──────────────────────────────────────────────────────────────────────────────

class SendFollowUpRequest(BaseModel):
    custom_message: Optional[str] = None  # Override the default template


class SendFollowUpResponse(BaseModel):
    success: bool
    message: str
    phone: Optional[str] = None


@router.post("/{work_order_id}/send-follow-up", response_model=SendFollowUpResponse)
async def send_follow_up_sms(
    work_order_id: str,
    request: SendFollowUpRequest,
    db: DbSession,
    current_user: CurrentUser,
):
    """Send a '$25 Off — Book Today' follow-up SMS to the customer on a draft/pending work order."""
    from app.services.sms_service import sms_service
    from app.models.message import Message

    # Load work order with customer
    from sqlalchemy.orm import selectinload
    result = await db.execute(
        select(WorkOrder)
        .options(selectinload(WorkOrder.customer))
        .where(WorkOrder.id == work_order_id)
    )
    wo = result.scalars().first()
    if not wo:
        raise HTTPException(status_code=404, detail="Work order not found")

    # Get customer phone
    customer = wo.customer
    phone = None
    if customer:
        phone = customer.phone or customer.mobile_phone
    if not phone:
        phone = wo.customer_phone
    if not phone:
        raise HTTPException(
            status_code=400,
            detail="No phone number on file for this customer. Cannot send follow-up.",
        )

    # Build the message
    customer_name = ""
    if customer:
        customer_name = customer.first_name or customer.last_name or ""
    if not customer_name and wo.customer_name:
        customer_name = wo.customer_name.split()[0] if wo.customer_name else ""

    total = wo.total_amount
    price_line = ""
    if total and float(total) > 0:
        price_line = f" Your quote: ${float(total):.2f}."

    if request.custom_message:
        body = request.custom_message
    else:
        body = (
            f"Hi {customer_name or 'there'}! Thanks for calling MAC Septic."
            f"{price_line}"
            f" Book today and save $25 — call us back or reply YES to schedule."
            f" Offer valid this week only."
        )

    # Send via RingCentral
    try:
        sms_response = await sms_service.send_sms(to=phone, body=body)
    except Exception as e:
        logger.error("Follow-up SMS failed", extra={"wo_id": work_order_id, "error": str(e)})
        raise HTTPException(status_code=500, detail="Failed to send follow-up SMS.")

    if sms_response.error:
        raise HTTPException(status_code=502, detail=f"SMS provider error: {sms_response.error[:200]}")

    # Record in messages table
    try:
        msg = Message(
            customer_id=wo.customer_id,
            work_order_id=wo.id,
            message_type="sms",
            direction="outbound",
            to_number=phone,
            content=body,
            status="sent",
            sent_at=datetime.utcnow(),
        )
        db.add(msg)

        # Mark WO as follow-up sent in checklist JSON
        checklist = wo.checklist or {}
        checklist["follow_up_sent"] = True
        checklist["follow_up_sent_at"] = datetime.utcnow().isoformat()
        wo.checklist = checklist

        await db.commit()
    except Exception:
        logger.warning("Failed to record follow-up message in DB", exc_info=True)

    return SendFollowUpResponse(
        success=True,
        message=f"Follow-up sent to {phone[-4:]}",
        phone=phone,
    )
