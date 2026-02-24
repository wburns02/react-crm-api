#!/usr/bin/env python3
"""Convert vehicle data to CRM asset import CSV.

Sources:
- ~/OneDrive/Documents/Luna Vehicles .xlsx (10 vehicles)
- ~/SharePoint/General/Annual Requirements.xlsx (TN/SC vehicles with VIN/tag/expiration)
"""

import csv
from pathlib import Path
import openpyxl

VEHICLES_SOURCE = Path.home() / "OneDrive/Documents/Luna Vehicles .xlsx"
COMPLIANCE_SOURCE = Path.home() / "SharePoint/General/Annual Requirements.xlsx"
OUTPUT = Path(__file__).parent / "assets_import.csv"

HEADERS = [
    "name", "asset_type", "asset_tag", "make", "model",
    "serial_number", "year", "purchase_price", "status",
    "condition", "odometer_miles", "notes"
]


def read_vehicles():
    """Read Luna Vehicles xlsx."""
    vehicles = []
    print(f"Reading {VEHICLES_SOURCE}...")
    wb = openpyxl.load_workbook(VEHICLES_SOURCE, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    # Find header row
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(c or "").lower() for c in row[:5]]
        if any("vehicle" in v or "make" in v or "model" in v or "year" in v for v in vals):
            header_row = i
            break

    if header_row is None:
        # Try row 0 as headers
        header_row = 0

    headers = [str(c or "").strip().lower() for c in rows[header_row]]
    print(f"  Found headers at row {header_row}: {headers}")

    for row in rows[header_row + 1:]:
        vals = dict(zip(headers, row))
        name_parts = []

        year = ""
        make = ""
        model = ""
        vin = ""
        plate = ""
        mileage = ""

        # Try common header names
        for k, v in vals.items():
            kl = k.lower()
            if "year" in kl:
                year = str(v or "").strip()
            elif "make" in kl:
                make = str(v or "").strip()
            elif "model" in kl:
                model = str(v or "").strip()
            elif "vin" in kl:
                vin = str(v or "").strip()
            elif "plate" in kl or "tag" in kl or "license" in kl:
                plate = str(v or "").strip()
            elif "mileage" in kl or "odometer" in kl or "miles" in kl:
                mileage = str(v or "").strip()
            elif "vehicle" in kl and not make:
                # Sometimes "Vehicle" col has combined description
                desc = str(v or "").strip()
                if desc:
                    name_parts.append(desc)

        name = " ".join(filter(None, [year, make, model])) or " ".join(name_parts)
        if not name or name.lower() in ("none", ""):
            continue

        vehicles.append({
            "name": name,
            "asset_type": "vehicle",
            "asset_tag": plate,
            "make": make,
            "model": model,
            "serial_number": vin,
            "year": year,
            "purchase_price": "",
            "status": "in_use",
            "condition": "good",
            "odometer_miles": mileage,
            "notes": f"License plate: {plate}" if plate else "Imported from Luna Vehicles"
        })

    wb.close()
    print(f"  Found {len(vehicles)} vehicles")
    return vehicles


def read_compliance_vehicles():
    """Read vehicle entries from Annual Requirements (fleet registrations)."""
    vehicles = []
    print(f"Reading {COMPLIANCE_SOURCE}...")
    wb = openpyxl.load_workbook(COMPLIANCE_SOURCE, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        req = str(row[0] or "").strip().lower()
        # Look for vehicle/fleet registration entries
        if any(term in req for term in ["vehicle", "fleet", "truck", "registration"]):
            name = str(row[0] or "").strip()
            permit = str(row[1] or "").strip()
            expiration = str(row[3] or "").strip()

            vehicles.append({
                "name": name,
                "asset_type": "vehicle",
                "asset_tag": permit,
                "make": "",
                "model": "",
                "serial_number": "",
                "year": "",
                "purchase_price": "",
                "status": "in_use",
                "condition": "good",
                "odometer_miles": "",
                "notes": f"Permit: {permit}, Expires: {expiration}" if permit else ""
            })

    wb.close()
    print(f"  Found {len(vehicles)} compliance vehicle entries")
    return vehicles


def main():
    vehicles = read_vehicles()
    compliance = read_compliance_vehicles()

    # Merge - dedup by VIN/name
    seen = set()
    all_vehicles = []
    for v in vehicles + compliance:
        key = v["serial_number"] or v["name"]
        if key.lower() not in seen:
            seen.add(key.lower())
            all_vehicles.append(v)

    with open(OUTPUT, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(all_vehicles)

    print(f"\nDone! Wrote {len(all_vehicles)} vehicles to {OUTPUT}")


if __name__ == "__main__":
    main()
